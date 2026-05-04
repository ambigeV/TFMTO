"""
plot_convergence_batch.py
-------------------------
Convergence plotter for the Data_Batch folder layout.

Data layout (algo-first, problem files flat inside each folder):
    DATA_ROOT/{algo_name}/{algo_name}_{prob_name}_{run_id}.pkl

Output mirrors plot_convergence.py:
    SAVE_ROOT/{prob_name}/{prob_name}_convergence.png
    one subplot per task, all algorithms overlaid, mean ± STD_SCALE*std shading.

Also generates a multi-sheet Excel results table (results_table_checkpoints.xlsx)
recording mean ± 0.5*std at iteration checkpoints 40/60/80/100, with Wilcoxon
rank-sum tests against the baseline and average ranks per checkpoint sheet.

Usage:
    python plot_convergence_batch.py                        # all problems, all methods
    python plot_convergence_batch.py -m BO MTBO             # filter to specific methods
    python plot_convergence_batch.py -m BO-TFM MTBO-TFM-Uni MTBO-TFM-Elite
    python plot_convergence_batch.py --table-only           # skip plots, generate table only
"""

import argparse
import pickle
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
from pathlib import Path

# =============================================================================
# Configuration — edit here
# =============================================================================

DATA_ROOT  = './Data/distill_demo'
SAVE_ROOT  = './Results_Batch_Demo'

# Problems to plot. Set to None to auto-discover all problems from pkl files,
# or list specific names e.g. ['P1', 'P3', 'P5'].
PROBLEMS: list | None = None
STD_SCALE  = 0.5
FIG_FORMAT = 'png'
DPI        = 150
ALPHA_BAND = 0.15

# Algorithms to include and their display order.
# Any algo found in the folder but not listed here is appended at the end.
ALGO_ORDER = [
    'GA', 'BO', 'BO-LCB', 'MTBO', 'BO-LCB-BCKT',
    'BO-TFM', 'MTBO-TFM-Uni', 'MTBO-TFM-Elite',
    'MTBO-TFM-Uni-OH', 'MTBO-TFM-Elite-OH',
    'BO-TFM-CMA', 'MTBO-TFM-Uni-CMA', 'MTBO-TFM-Elite-CMA',
    'MTBO-TFM-Uni-OH-CMA', 'MTBO-TFM-Elite-OH-CMA',
]

# --- Table generation config ---
# Iteration checkpoints to snapshot in the results table.
CHECKPOINTS: list[int] = [40, 60, 80, 100]

# Baseline algorithm for Wilcoxon +/-/= comparison.
# None → last algorithm in the final ordered list is used as baseline.
BASELINE_ALGO: str | None = None

# p-value threshold for the rank-sum test.
TABLE_SIGNIFICANCE: float = 0.05

# Std multiplier for table display (half-std).
TABLE_STD_SCALE: float = 0.5

MARKERS = ['o', 's', '^', 'D', 'v', 'P', 'X', '*', 'h', '<', '>', 'p', 'H', '8', '+']


def make_color_list(n: int) -> list:
    """Return n visually distinct colors using tab20 for ≤20, then hsv for more."""
    import matplotlib.cm as cm
    if n <= 20:
        cmap = cm.get_cmap('tab20', n)
    else:
        cmap = cm.get_cmap('hsv', n)
    return [cmap(i) for i in range(n)]


# =============================================================================
# Helpers
# =============================================================================

def load_pkl(path: Path) -> dict:
    with open(path, 'rb') as f:
        return pickle.load(f)


def best_so_far(all_objs_task: list) -> np.ndarray:
    """Staircase all_objs[task] → monotone best-so-far curve."""
    curve = np.array([np.min(gen[:, 0]) for gen in all_objs_task])
    for i in range(1, len(curve)):
        curve[i] = min(curve[i], curve[i - 1])
    return curve


def parse_prob_name(stem: str, algo_name: str) -> str:
    """
    Extract the problem name from a pkl stem.

    Stem format:  {algo_name}_{prob_name}_{run_id}
    e.g.  'MTBO-TFM-Uni_P3_2'  with algo_name='MTBO-TFM-Uni'  → 'P3'

    Works regardless of hyphens/underscores in algo_name because we strip
    the known prefix (algo_name + '_') and then split off the trailing run id.
    """
    suffix = stem[len(algo_name) + 1:]   # '{prob_name}_{run_id}'
    prob, _ = suffix.rsplit('_', 1)
    return prob


def discover_problems(data_root: Path, algo_names: list) -> list:
    """Scan all pkl files and collect unique problem names."""
    problems = set()
    for algo in algo_names:
        algo_dir = data_root / algo
        if not algo_dir.exists():
            continue
        for pkl in algo_dir.glob('*.pkl'):
            try:
                prob = parse_prob_name(pkl.stem, algo)
                problems.add(prob)
            except (ValueError, IndexError):
                pass
    return sorted(problems)


def load_algo_data(data_root: Path, algo_name: str, prob_name: str):
    """
    Load all run pkl files for (algo_name, prob_name) from the batch layout.

    Returns
    -------
    curves_per_task  : list[list[np.ndarray]]  — [task][run] = convergence array
    max_nfes_per_task: list[int]
    n_tasks          : int
    None, None, 0 if no data found.
    """
    algo_dir  = data_root / algo_name
    if not algo_dir.exists():
        return None, None, 0

    pkl_files = sorted(algo_dir.glob(f'{algo_name}_{prob_name}_*.pkl'))
    if not pkl_files:
        return None, None, 0

    all_runs          = []
    max_nfes_per_task = None

    for pkl_path in pkl_files:
        try:
            data = load_pkl(pkl_path)
        except (EOFError, pickle.UnpicklingError, Exception) as e:
            print(f'  [WARN] Skipping corrupt/incomplete file: {pkl_path.name} ({type(e).__name__}: {e})')
            continue
        all_objs  = data['all_objs']
        run_curves = [best_so_far(all_objs[t]) for t in range(len(all_objs))]
        all_runs.append(run_curves)
        if max_nfes_per_task is None:
            max_nfes_per_task = list(data['max_nfes'])

    if not all_runs:
        return None, None, 0

    n_tasks = len(all_runs[0])
    curves_per_task = [
        [all_runs[r][t] for r in range(len(all_runs))]
        for t in range(n_tasks)
    ]
    return curves_per_task, max_nfes_per_task, n_tasks


def align_curves(curves: list) -> np.ndarray:
    """Truncate curves to the shortest length and stack → (n_runs, min_len)."""
    min_len = min(len(c) for c in curves)
    return np.stack([c[:min_len] for c in curves], axis=0)


# =============================================================================
# Main plot function
# =============================================================================

def plot_problem(prob_name: str, data_root: Path, save_root: Path, algos: list):
    save_dir = save_root / prob_name
    save_dir.mkdir(parents=True, exist_ok=True)

    # --- load data for every algorithm that has files for this problem ---
    algo_data     = {}
    algo_max_nfes = {}
    n_tasks       = 0

    for algo in algos:
        curves, max_nfes_pt, nt = load_algo_data(data_root, algo, prob_name)
        if curves is None:
            continue
        algo_data[algo]     = curves
        algo_max_nfes[algo] = max_nfes_pt
        n_tasks = max(n_tasks, nt)

    if not algo_data:
        print(f'  [SKIP] {prob_name}: no data found for any algorithm.')
        return

    n_runs_info = {a: len(algo_data[a][0]) for a in algo_data}
    print(f'  {prob_name}: {len(algo_data)} algos, {n_tasks} tasks — '
          f'runs: { {a: n for a, n in n_runs_info.items()} }')

    # --- colour / marker maps (stable across problems) ---
    all_algo_keys = list(algo_data.keys())
    ordered_algos = ALGO_ORDER + [x for x in all_algo_keys if x not in ALGO_ORDER]
    colors = make_color_list(len(ordered_algos))
    color_map  = {}
    marker_map = {}
    for i, a in enumerate(ordered_algos):
        color_map[a]  = colors[i]
        marker_map[a] = MARKERS[i % len(MARKERS)]

    # --- figure ---
    fig, axes = plt.subplots(1, n_tasks, figsize=(5 * n_tasks, 4), squeeze=False)
    fig.suptitle(f'CEC17-MTSO-10D  —  {prob_name}', fontsize=13, fontweight='bold')

    for t in range(n_tasks):
        ax = axes[0][t]
        ax.set_title(f'Task {t + 1}', fontsize=11)
        ax.set_xlabel('NFEs', fontsize=10)
        ax.set_ylabel('Best Objective Found', fontsize=10)

        for algo, curves_per_task in algo_data.items():
            if t >= len(curves_per_task):
                continue

            mat  = align_curves(curves_per_task[t])   # (n_runs, n_pts)
            mean = mat.mean(axis=0)
            std  = mat.std(axis=0)

            nfes       = algo_max_nfes[algo][t] if t < len(algo_max_nfes[algo]) else len(mean)
            x          = np.linspace(0, nfes, len(mean))
            color      = color_map.get(algo, '#333333')
            marker     = marker_map.get(algo, 'o')
            mark_every = max(1, len(x) // 10)

            ax.plot(x, mean,
                    label=algo,
                    color=color,
                    marker=marker,
                    markevery=mark_every,
                    markersize=4,
                    linewidth=1.5)

            if mat.shape[0] > 1:   # shade only when multiple runs exist
                ax.fill_between(
                    x,
                    mean - STD_SCALE * std,
                    mean + STD_SCALE * std,
                    alpha=ALPHA_BAND,
                    color=color,
                )

        ax.xaxis.set_major_locator(ticker.MaxNLocator(integer=True, nbins=6))
        ax.grid(True, linestyle='--', alpha=0.4)

    handles, labels = axes[0][0].get_legend_handles_labels()
    fig.legend(
        handles, labels,
        loc='lower center',
        ncol=min(len(algo_data), 5),
        fontsize=9,
        bbox_to_anchor=(0.5, -0.12),
        frameon=True,
    )

    plt.tight_layout()
    out_path = save_dir / f'{prob_name}_convergence.{FIG_FORMAT}'
    fig.savefig(out_path, dpi=DPI, bbox_inches='tight')
    plt.close(fig)
    print(f'    Saved → {out_path}')


# =============================================================================
# Results table generation
# =============================================================================

def _assign_ranks(means: dict) -> dict:
    """
    Assign average ranks (1 = best / lowest) to a dict of {algo: mean_value}.
    Ties receive equal average rank.
    """
    sorted_algos = sorted(means, key=lambda a: means[a])
    ranks = {}
    i = 0
    while i < len(sorted_algos):
        j = i + 1
        while j < len(sorted_algos) and means[sorted_algos[j]] == means[sorted_algos[i]]:
            j += 1
        avg_r = (i + 1 + j) / 2.0
        for k in range(i, j):
            ranks[sorted_algos[k]] = avg_r
        i = j
    return ranks


def _parse_mean(cell_str: str) -> float | None:
    """Parse the mean float from a formatted cell string like '1.23e+00 (4.5e-02) +'."""
    try:
        return float(cell_str.split('(')[0].strip())
    except Exception:
        return None


def _format_results_wb(wb, algos: list) -> None:
    """Apply uniform formatting to every sheet in the workbook."""
    try:
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    except ImportError:
        return

    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin'),
    )
    normal_font = Font(name='Times New Roman', size=10)
    bold_font   = Font(name='Times New Roman', size=10, bold=True)
    green_fill  = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    gray_fill   = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    center_align = Alignment(horizontal='center', vertical='center')

    # Column indices of algo cells (1-based): Problem=1, Task=2, algos start at 3
    algo_col_start = 3

    for ws in wb.worksheets:
        max_row = ws.max_row
        max_col = ws.max_column

        # Identify summary rows (last two rows: +/-/= and Average Rank)
        summary_rows = {max_row - 1, max_row}

        for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
            for cell in row:
                cell.border = thin
                cell.alignment = center_align
                r = cell.row
                if r == 1 or r in summary_rows:
                    cell.font = bold_font
                else:
                    cell.font = normal_font
                if r in summary_rows:
                    cell.fill = gray_fill

        # Highlight best (lowest mean) algo cell per data row
        for r in range(2, max_row - 1):  # data rows only
            best_mean = None
            best_col  = None
            for c in range(algo_col_start, max_col + 1):
                val = ws.cell(row=r, column=c).value
                if val and isinstance(val, str):
                    m = _parse_mean(val)
                    if m is not None and (best_mean is None or m < best_mean):
                        best_mean = m
                        best_col  = c
            if best_col is not None:
                ws.cell(row=r, column=best_col).fill = green_fill

        # Auto-adjust column widths
        for col in ws.columns:
            max_len = max(
                (len(str(cell.value)) for cell in col if cell.value is not None),
                default=8,
            )
            col[0].parent.column_dimensions[col[0].column_letter].width = min(max_len + 2, 32)


def generate_results_table(
    data_root: Path,
    save_root: Path,
    algos: list,
    problems: list,
    checkpoints: list | None = None,
    baseline_algo: str | None = None,
    sig_level: float = TABLE_SIGNIFICANCE,
    std_scale: float = TABLE_STD_SCALE,
) -> None:
    """
    Generate a multi-sheet Excel table recording best-so-far performance at
    specified iteration checkpoints.

    Each sheet (Iter40, Iter60, Iter80, Iter100) contains:
      - mean ± std_scale*std per (problem, task) for every algorithm
      - Wilcoxon rank-sum '+'/'-'/'=' symbols vs the baseline algorithm
      - '+/-/=' summary row and Average Rank row

    Parameters
    ----------
    data_root   : Path to the batch data folder.
    save_root   : Path to the output folder (xlsx saved here).
    algos       : Ordered list of algorithm names to include.
    problems    : Ordered list of problem names.
    checkpoints : Iteration indices to snapshot (1-based, e.g. [40, 60, 80, 100]).
    baseline_algo : Algorithm to compare against; defaults to last in algos.
    sig_level   : Significance level for Wilcoxon rank-sum test.
    std_scale   : Multiplier applied to std in display (0.5 = half-std).
    """
    try:
        from scipy import stats as sp_stats
        from openpyxl import Workbook
    except ImportError as e:
        print(f'[TABLE] Missing dependency: {e}. Skipping table generation.')
        return

    if checkpoints is None:
        checkpoints = CHECKPOINTS

    baseline = baseline_algo if (baseline_algo and baseline_algo in algos) else algos[-1]
    non_base = [a for a in algos if a != baseline]

    print(f'\n[TABLE] Baseline: {baseline}')
    print(f'[TABLE] Checkpoints: {checkpoints}')

    # Load all data once
    print('[TABLE] Loading data...')
    all_data: dict = {}   # all_data[algo][prob] = (curves_per_task, n_tasks)
    for algo in algos:
        all_data[algo] = {}
        for prob in problems:
            curves, _, nt = load_algo_data(data_root, algo, prob)
            if curves is not None:
                all_data[algo][prob] = (curves, nt)

    wb = Workbook()
    wb.remove(wb.active)  # remove the blank default sheet

    for ck in checkpoints:
        ck_idx = ck - 1   # convert 1-based iteration to 0-based array index
        ws = wb.create_sheet(title=f'Iter{ck}')

        # Write header
        header = ['Problem', 'Task'] + algos
        for c, h in enumerate(header, 1):
            ws.cell(row=1, column=c, value=h)

        # Accumulate per-sheet stats
        comp_counts = {a: [0, 0, 0] for a in non_base}   # [+, -, =]
        algo_ranks: dict[str, list] = {a: [] for a in algos}
        data_row = 2

        for prob in problems:
            n_tasks = max(
                (all_data[a][prob][1] for a in algos if prob in all_data.get(a, {}) and prob in all_data[a]),
                default=0,
            )

            for t in range(n_tasks):
                # Extract checkpoint value for each algo × run
                algo_vals: dict[str, np.ndarray] = {}
                for algo in algos:
                    entry = all_data.get(algo, {}).get(prob)
                    if entry is None:
                        continue
                    curves_pt, nt = entry
                    if t >= nt or not curves_pt[t]:
                        continue
                    mat = align_curves(curves_pt[t])          # (n_runs, n_pts)
                    col = min(ck_idx, mat.shape[1] - 1)       # clamp if curve shorter
                    algo_vals[algo] = mat[:, col]

                # Assign per-row ranks (lower mean = better rank)
                means = {a: float(np.mean(v)) for a, v in algo_vals.items()}
                row_ranks = _assign_ranks(means) if means else {}
                for a in algos:
                    algo_ranks[a].append(row_ranks.get(a, np.nan))

                # Write Problem and Task columns
                ws.cell(row=data_row, column=1, value=prob)
                ws.cell(row=data_row, column=2, value=t + 1)

                base_vals = algo_vals.get(baseline, np.array([]))

                for ci, algo in enumerate(algos, 3):
                    if algo not in algo_vals:
                        ws.cell(row=data_row, column=ci, value='N/A')
                        continue

                    v = algo_vals[algo]
                    mean = float(np.mean(v))
                    std  = float(np.std(v, ddof=1)) if len(v) > 1 else 0.0
                    cell_str = f'{mean:.3e} ({std_scale * std:.1e})'

                    if algo != baseline and len(base_vals) > 0:
                        try:
                            v_h = np.mean(v) + 0.5 * (v - np.mean(v))
                            b_h = np.mean(base_vals) + 0.5 * (base_vals - np.mean(base_vals))
                            _, pval = sp_stats.ranksums(v_h, b_h)
                            if pval < sig_level:
                                sym = '+' if np.mean(v) < np.mean(base_vals) else '-'
                            else:
                                sym = '='
                        except Exception:
                            sym = '='
                        if   sym == '+': comp_counts[algo][0] += 1
                        elif sym == '-': comp_counts[algo][1] += 1
                        else:            comp_counts[algo][2] += 1
                        ws.cell(row=data_row, column=ci, value=f'{cell_str} {sym}')
                    else:
                        ws.cell(row=data_row, column=ci, value=cell_str)

                data_row += 1

        # +/-/= summary row
        ws.cell(row=data_row, column=1, value='+/-/=')
        for ci, algo in enumerate(algos, 3):
            if algo == baseline:
                ws.cell(row=data_row, column=ci, value='Base')
            else:
                c = comp_counts[algo]
                ws.cell(row=data_row, column=ci, value=f'{c[0]}/{c[1]}/{c[2]}')
        data_row += 1

        # Average Rank row — NaN rows (missing data) are treated as last place + 1
        # so that averages are computed over the same number of rows for all algos.
        n_rows = data_row - 2  # total data rows written
        n_algos_present = len([a for a in algos if any(not np.isnan(r) for r in algo_ranks[a])])
        worst_rank = n_algos_present + 1
        ws.cell(row=data_row, column=1, value='Average Rank')
        for ci, algo in enumerate(algos, 3):
            filled = [r if not np.isnan(r) else worst_rank for r in algo_ranks[algo]]
            ws.cell(row=data_row, column=ci, value=f'{np.mean(filled):.2f}' if filled else 'N/A')

        print(f'  [TABLE] Sheet Iter{ck}: {data_row - 2} data rows written.')

    # Apply formatting and save
    _format_results_wb(wb, algos)
    save_root.mkdir(parents=True, exist_ok=True)
    out = save_root / 'results_table_checkpoints.xlsx'
    wb.save(out)
    print(f'[TABLE] Saved → {out}')


# =============================================================================
# Entry point
# =============================================================================

if __name__ == '__main__':
    parser = argparse.ArgumentParser(
        description='Plot convergence curves and generate results table from batch data.'
    )
    parser.add_argument(
        '-m', '--methods',
        nargs='+',
        metavar='METHOD',
        default=None,
        help='Methods/algorithms to include (e.g. -m BO MTBO BO-TFM). '
             'Defaults to all algorithm folders found on disk.',
    )
    parser.add_argument(
        '--table-only',
        action='store_true',
        help='Skip convergence plots; only generate the results table.',
    )
    parser.add_argument(
        '--no-table',
        action='store_true',
        help='Skip results table generation; only produce convergence plots.',
    )
    parser.add_argument(
        '--checkpoints',
        nargs='+',
        type=int,
        metavar='ITER',
        default=None,
        help='Iteration checkpoints for the results table (e.g. --checkpoints 40 60 80 100).',
    )
    parser.add_argument(
        '--baseline',
        metavar='ALGO',
        default=None,
        help='Baseline algorithm for Wilcoxon +/-/= comparison. '
             'Defaults to the last algorithm in the ordered list.',
    )
    args = parser.parse_args()

    data_root = Path(DATA_ROOT)
    save_root = Path(SAVE_ROOT)

    if not data_root.exists():
        print(f'ERROR: Data_Batch folder not found: {data_root.resolve()}')
        raise SystemExit(1)

    # Collect algo folders present on disk (preserve ALGO_ORDER, append extras)
    present_algos = {d.name for d in data_root.iterdir() if d.is_dir()}

    if args.methods is not None:
        missing = [m for m in args.methods if m not in present_algos]
        if missing:
            print(f'[WARN] Requested methods not found on disk: {missing}')
        algos = [a for a in args.methods if a in present_algos]
    else:
        algos = [a for a in ALGO_ORDER if a in present_algos] + \
                sorted(present_algos - set(ALGO_ORDER))

    if not algos:
        print(f'ERROR: No algorithm folders found in {data_root.resolve()}')
        raise SystemExit(1)

    # Problems: from script-level config or auto-discovered from file names
    if PROBLEMS is not None:
        problems = PROBLEMS
    else:
        problems = discover_problems(data_root, algos)
        if not problems:
            print('ERROR: Could not discover any problem names from pkl files.')
            raise SystemExit(1)
        print(f'Auto-discovered problems: {problems}')

    print(f'Algorithms to plot: {algos}')
    print(f'Problems to plot:   {problems}\n')

    # --- Convergence plots ---
    if not args.table_only:
        for prob in problems:
            plot_problem(prob, data_root, save_root, algos)
        print('\nPlots done.')

    # --- Results table ---
    if not args.no_table:
        checkpoints = args.checkpoints if args.checkpoints else CHECKPOINTS
        baseline    = args.baseline    if args.baseline    else BASELINE_ALGO
        generate_results_table(
            data_root    = data_root,
            save_root    = save_root,
            algos        = algos,
            problems     = problems,
            checkpoints  = checkpoints,
            baseline_algo= baseline,
        )

    print('\nAll done.')
