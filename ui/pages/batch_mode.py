"""Batch Experiment Mode - Multiple runs with statistical analysis."""

import os
import io
import time
import threading
import contextlib
from pathlib import Path
import numpy as np
import dearpygui.dearpygui as dpg

from utils.registry import (
    get_algorithm_names, get_algorithm_class,
    get_problem_suites, get_problem_methods, get_problem_creator,
    is_multi_objective, get_problem_settings, get_problem_module_path,
)
from utils.algo_scanner import (
    scan_all_algorithms,
    get_algorithm_params_from_scan,
    get_common_params,
    get_algorithm_info,
    format_algorithm_info,
    get_discovered_algorithm_module_info,
)
from utils.problem_scanner import (
    get_scanned_problem_params, is_fixed_dimension_problem,
)
from utils.runner import count_pkl_files, BatchStatus, get_default_workers
from utils.file_manager import FileManager
from utils.backup_manager import BackupManager
from config.default_params import PROBLEM_PARAMS, FIXED_DIMENSION_SUITES, FIXED_OBJECTIVES_SUITES
from config.constants import (
    CATEGORIES, ALGO_CATEGORIES, METRICS, TABLE_FORMATS, FIGURE_FORMATS, STATISTIC_TYPES,
    COLOR_TITLE, COLOR_ERROR, COLOR_SUCCESS, COLOR_SECTION
)
from components.dpg_helpers import (
    add_checkbox_group, get_checkbox_selections, update_checkbox_group,
    set_all_checkboxes, load_image_to_texture, show_error_modal, show_info_modal,
    show_confirm_dialog, format_time, copy_text_to_clipboard, _on_copy_error_click,
    get_disabled_btn_theme as _get_disabled_btn_theme,
    parse_param_value as _parse_param_value,
    open_file_location as _open_file_location,
    copy_image_to_clipboard as _copy_image_to_clipboard,
    open_algo_source as _open_algo_source,
)

# Module state
_state = {
    "status": None,
    "file_manager": None,
    "backup_manager": None,
    "last_poll": 0.0,
    "displayed": False,
    "load_data_save_path": None,
    "current_prob_category": "STSO",
    "current_algo_category": "STSO",
    "current_suite": "",
    "selected_probs": [],  # List of selected problems: [(suite, method), ...]
    "prob_names": {},  # {index: custom_name} for renamed problems
    "prob_params": {},  # {index: {param: value}} for problem parameters
    "selected_algos": [],  # List of selected algorithms
    "algo_names": {},  # {index: custom_name} for renamed algorithms
    "algo_param_widgets": {},
    "stdout_buffer": "",  # Captured stdout from worker thread
}


def _init_managers(base_path: str):
    """Initialize file and backup managers."""
    _state["file_manager"] = FileManager(base_path)
    _state["backup_manager"] = BackupManager(base_path)
    _state["file_manager"].ensure_structure()
    # Pre-scan algorithm parameters
    scan_all_algorithms()


def _on_prob_category_change(sender, app_data):
    """Update suite combo when problem category changes."""
    cat = app_data
    _state["current_prob_category"] = cat
    suites = get_problem_suites(cat)
    dpg.configure_item("batch_suite_combo", items=suites)
    if suites:
        dpg.set_value("batch_suite_combo", suites[0])
        _on_suite_change(None, suites[0])
    _update_metric_visibility(cat)


def _on_algo_category_change(sender, app_data):
    """Update algorithm list when algorithm category changes."""
    cat = app_data
    _state["current_algo_category"] = cat
    _update_algo_list_display()


def _on_suite_change(sender, app_data):
    """Update problems list when suite changes."""
    suite = app_data
    _state["current_suite"] = suite
    _update_prob_list_display()


def _update_metric_visibility(cat: str):
    """No-op: metric combo is always visible now."""
    pass


def _on_prob_click(sender, app_data, user_data):
    """Handle problem button click to add to selected list."""
    suite, method = user_data
    _state["selected_probs"].append((suite, method))
    _update_prob_params_display()


def _remove_prob(sender, app_data, user_data):
    """Remove problem from selected list by index."""
    idx = user_data
    if 0 <= idx < len(_state["selected_probs"]):
        _state["selected_probs"].pop(idx)
        # Clean up params for this index
        if idx in _state["prob_names"]:
            del _state["prob_names"][idx]
        if idx in _state["prob_params"]:
            del _state["prob_params"][idx]
        _update_prob_params_display()


def _move_prob_up(sender, app_data, user_data):
    """Move problem up in the selected list."""
    idx = user_data
    selected = _state["selected_probs"]
    if idx <= 0 or idx >= len(selected):
        return
    selected[idx], selected[idx - 1] = selected[idx - 1], selected[idx]
    # Swap names and params
    names = _state["prob_names"]
    params = _state["prob_params"]
    names[idx], names[idx - 1] = names.get(idx - 1), names.get(idx)
    params[idx], params[idx - 1] = params.get(idx - 1, {}), params.get(idx, {})
    _update_prob_params_display()


def _move_prob_down(sender, app_data, user_data):
    """Move problem down in the selected list."""
    idx = user_data
    selected = _state["selected_probs"]
    if idx < 0 or idx >= len(selected) - 1:
        return
    selected[idx], selected[idx + 1] = selected[idx + 1], selected[idx]
    # Swap names and params
    names = _state["prob_names"]
    params = _state["prob_params"]
    names[idx], names[idx + 1] = names.get(idx + 1), names.get(idx)
    params[idx], params[idx + 1] = params.get(idx + 1, {}), params.get(idx, {})
    _update_prob_params_display()


def _update_prob_list_display():
    """Update the problem list with clickable buttons."""
    if not dpg.does_item_exist("batch_prob_list_container"):
        return

    dpg.delete_item("batch_prob_list_container", children_only=True)

    cat = _state["current_prob_category"]
    suite = _state["current_suite"]
    if not suite:
        return

    methods = get_problem_methods(cat, suite)
    for method in methods:
        dpg.add_button(label=method, callback=_on_prob_click, user_data=(suite, method),
                       width=-1, parent="batch_prob_list_container")


def _on_prob_name_change(idx: int, new_name: str):
    """Handle problem name change."""
    _state["prob_names"][idx] = new_name


def _on_prob_param_change(idx: int, param: str, value):
    """Handle problem parameter change."""
    if idx not in _state["prob_params"]:
        _state["prob_params"][idx] = {}
    _state["prob_params"][idx][param] = value


def _get_prob_params_for_suite(suite: str) -> dict:
    """Get available parameters for a problem suite (auto-scanned or hardcoded)."""
    params = {}
    cat = _state["current_prob_category"]

    # Try auto-scanned params first
    scanned_params = get_scanned_problem_params(cat, suite)
    suite_params = scanned_params if scanned_params else PROBLEM_PARAMS.get(suite, {})

    # Use scanner to determine if fixed dimension
    is_fixed_dim = is_fixed_dimension_problem(cat, suite) if scanned_params else (suite in FIXED_DIMENSION_SUITES)

    # Check if objectives are fixed (e.g., ZDT, CF, UF have fixed M per problem)
    is_fixed_obj = suite in FIXED_OBJECTIVES_SUITES

    # Check if suite needs D (decision variables)
    if not is_fixed_dim:
        if "D" in suite_params:
            params["D"] = {"type": "int", "default": suite_params["D"].get("default", 50)}
        else:
            params["D"] = {"type": "int", "default": 50}

    # Check other parameters - M only if not fixed-objective suite
    if "M" in suite_params and not is_fixed_obj:
        params["M"] = {"type": "int", "default": suite_params["M"].get("default", 3)}
    if "K" in suite_params:
        params["K"] = {"type": "int", "default": suite_params["K"].get("default", 10)}
    if "L" in suite_params:
        params["L"] = {"type": "int", "default": suite_params["L"].get("default", 20)}

    return params


def _update_prob_params_display():
    """Update problem parameters display based on selection."""
    if not dpg.does_item_exist("batch_prob_params_container"):
        return

    dpg.delete_item("batch_prob_params_container", children_only=True)

    selected = _state["selected_probs"]
    if not selected:
        dpg.add_text("Click problems above to add",
                     parent="batch_prob_params_container", color=(150, 150, 150))
        return

    for i, (suite, method) in enumerate(selected):
        # Get custom name if set
        default_name = method
        custom_name = _state["prob_names"].get(i, default_name)
        header_label = custom_name if custom_name != default_name else default_name

        # Up/Down buttons
        with dpg.group(horizontal=True, parent="batch_prob_params_container"):
            dpg.add_button(label="^", width=20, callback=_move_prob_up, user_data=i,
                          enabled=(i > 0))
            dpg.add_button(label="v", width=20, callback=_move_prob_down, user_data=i,
                          enabled=(i < len(selected) - 1))

        # Collapsing header
        with dpg.collapsing_header(label=f"{header_label}##{suite}_{method}_{i}",
                                   parent="batch_prob_params_container",
                                   default_open=(len(selected) == 1)):
            # Name parameter
            with dpg.group(horizontal=True):
                dpg.add_text("Name:")
                name_tag = f"batch_prob_name_{suite}_{method}_{i}"
                dpg.add_input_text(tag=name_tag, default_value=custom_name, width=-1,
                                   callback=lambda s, a, u: _on_prob_name_change(u, a),
                                   user_data=i)

            # Problem parameters based on suite
            params = _get_prob_params_for_suite(suite)
            saved_params = _state["prob_params"].get(i, {})

            for param_name, param_info in params.items():
                with dpg.group(horizontal=True):
                    dpg.add_text(f"{param_name}:")
                    tag = f"batch_prob_param_{suite}_{method}_{i}_{param_name}"
                    default_val = saved_params.get(param_name, param_info['default'])
                    if default_val is None:
                        default_val = 50 if param_name == "D" else 3 if param_name == "M" else 10
                    dpg.add_input_int(tag=tag, default_value=int(default_val), width=-1, step=0,
                                      callback=lambda s, a, u: _on_prob_param_change(u[0], u[1], a),
                                      user_data=(i, param_name))

            # Delete button
            dpg.add_spacer(height=5)
            del_btn = dpg.add_button(label="Delete Problem", callback=_remove_prob, user_data=i, width=-1)
            if _state.get("delete_theme"):
                dpg.bind_item_theme(del_btn, _state["delete_theme"])


def _on_algo_click(sender, app_data, user_data):
    """Handle algorithm button click to add to selected list."""
    algo_name = user_data
    _state["selected_algos"].append(algo_name)
    _update_algo_params_display()


def _remove_algo(sender, app_data, user_data):
    """Remove algorithm from selected list by index."""
    idx = user_data
    if 0 <= idx < len(_state["selected_algos"]):
        _state["selected_algos"].pop(idx)
        _update_algo_params_display()


def _update_algo_list_display():
    """Update the algorithm list with clickable buttons and tooltips."""
    if not dpg.does_item_exist("batch_algo_list_container"):
        return

    dpg.delete_item("batch_algo_list_container", children_only=True)

    cat = _state["current_algo_category"]
    algos = get_algorithm_names(cat)

    # Labels for algorithm info
    info_labels = {
        'n_tasks': 'Tasks', 'dims': 'Dimensions', 'objs': 'Objectives',
        'n_objs': 'Num Objectives', 'cons': 'Constraints', 'n_cons': 'Num Constraints',
        'expensive': 'Expensive', 'knowledge_transfer': 'Knowledge Transfer',
        'n': 'Population', 'max_nfes': 'Max NFEs',
    }

    for algo_name in algos:
        btn = dpg.add_button(label=algo_name, callback=_on_algo_click, user_data=algo_name,
                             width=-1, parent="batch_algo_list_container")
        # Add tooltip with algorithm information (blue labels)
        info = get_algorithm_info(cat, algo_name)
        if info:
            with dpg.tooltip(btn):
                for key, value in info.items():
                    label = info_labels.get(key, key)
                    with dpg.group(horizontal=True):
                        dpg.add_text(f"{label}:", color=COLOR_TITLE)
                        dpg.add_text(f" {value}")


def _clear_algo_params():
    """Clear algorithm parameter widgets."""
    if dpg.does_item_exist("batch_algo_params_container"):
        dpg.delete_item("batch_algo_params_container", children_only=True)
    _state["algo_param_widgets"] = {}


def _move_algo_up(sender, app_data, user_data):
    """Move algorithm up in the selected list."""
    idx = user_data
    selected = _state["selected_algos"]
    if idx <= 0 or idx >= len(selected):
        return
    selected[idx], selected[idx - 1] = selected[idx - 1], selected[idx]
    # Also swap custom names if they exist
    names = _state["algo_names"]
    if idx in names or (idx - 1) in names:
        names[idx], names[idx - 1] = names.get(idx - 1, selected[idx]), names.get(idx, selected[idx - 1])
    _update_algo_params_display()


def _move_algo_down(sender, app_data, user_data):
    """Move algorithm down in the selected list."""
    idx = user_data
    selected = _state["selected_algos"]
    if idx < 0 or idx >= len(selected) - 1:
        return
    selected[idx], selected[idx + 1] = selected[idx + 1], selected[idx]
    # Also swap custom names if they exist
    names = _state["algo_names"]
    if idx in names or (idx + 1) in names:
        names[idx], names[idx + 1] = names.get(idx + 1, selected[idx]), names.get(idx, selected[idx + 1])
    _update_algo_params_display()


def _update_algo_params_display():
    """Update algorithm parameters display based on selection."""
    _clear_algo_params()

    if not dpg.does_item_exist("batch_algo_params_container"):
        return

    selected = _state["selected_algos"]
    cat = _state["current_algo_category"]

    if not selected:
        dpg.add_text("Click algorithms above to add",
                     parent="batch_algo_params_container", color=(150, 150, 150))
        return

    # Each algorithm as a collapsible header with X button
    for i, algo_name in enumerate(selected):
        params = get_algorithm_params_from_scan(cat, algo_name)
        safe_algo_name = algo_name.replace("/", "-")

        # Get custom name if set, otherwise use original
        custom_name = _state["algo_names"].get(i, algo_name)
        header_label = custom_name if custom_name != algo_name else algo_name

        # Collapsing header row with algorithm name and controls
        with dpg.group(horizontal=True, parent="batch_algo_params_container"):
            # Up/Down buttons for reordering
            dpg.add_button(label="^", width=20, callback=_move_algo_up, user_data=i,
                          enabled=(i > 0))
            dpg.add_button(label="v", width=20, callback=_move_algo_down, user_data=i,
                          enabled=(i < len(selected) - 1))

        # Collapsing header with algorithm name
        with dpg.collapsing_header(label=f"{header_label}##{safe_algo_name}_{i}",
                                   parent="batch_algo_params_container",
                                   default_open=(len(selected) == 1)) as header:
            # Right-click popup menu on header
            with dpg.popup(header, mousebutton=dpg.mvMouseButton_Right,
                           min_size=(160, 0), max_size=(160, 50)):
                dpg.add_menu_item(label="Open Source File",
                                  callback=_open_algo_source, user_data=(algo_name, cat))

            # Name parameter for renaming
            with dpg.group(horizontal=True):
                dpg.add_text("Name:")
                name_tag = f"batch_algo_name_{safe_algo_name}_{i}"
                dpg.add_input_text(tag=name_tag, default_value=custom_name, width=-1,
                                   callback=lambda s, a, u: _on_algo_name_change(u[0], a),
                                   user_data=(i,))

            # Other parameters
            for param_name, param_info in params.items():
                _add_algo_param_input(algo_name, param_name, param_info, i)

            # Delete button at the end
            dpg.add_spacer(height=5)
            del_btn = dpg.add_button(label="Delete Algorithm", callback=_remove_algo, user_data=i,
                          width=-1)
            if _state.get("delete_theme"):
                dpg.bind_item_theme(del_btn, _state["delete_theme"])


def _on_algo_name_change(idx: int, new_name: str):
    """Handle algorithm name change."""
    _state["algo_names"][idx] = new_name


def _add_algo_param_input(algo_name: str, param_name: str, param_info: dict, idx: int):
    """Add a parameter input widget for an algorithm."""
    param_type = param_info['type']
    default = param_info['default']

    # Use safe name and index for unique tags (replace "/" with "-")
    safe_algo_name = algo_name.replace("/", "-")
    tag = f"batch_algo_param_{safe_algo_name}_{idx}_{param_name}"
    _state["algo_param_widgets"][f"{safe_algo_name}_{idx}__{param_name}"] = tag

    # Check if parameter is vectorizable based on algorithm_information
    cat = _state["current_algo_category"]
    algo_info = get_algorithm_info(cat, algo_name)
    is_vectorizable = algo_info.get(param_name, '') == 'unequal'

    with dpg.group(horizontal=True):
        # Add [list] hint for vectorizable parameters
        if is_vectorizable:
            dpg.add_text(f"{param_name}:", color=COLOR_SECTION)
            # Use text input for vectorizable parameters to allow list input
            default_str = str(default) if default is not None else "100" if param_type == 'int' else "0.5"
            dpg.add_input_text(tag=tag, default_value=default_str, width=-1, hint="e.g. [100,200]")
        else:
            dpg.add_text(f"{param_name}:")
            if param_type == 'int':
                dpg.add_input_int(tag=tag, default_value=default if default is not None else 100,
                                  width=-1, step=0)
            elif param_type == 'float':
                dpg.add_input_float(tag=tag, default_value=default if default is not None else 0.5,
                                    width=-1, step=0, format="%.4f")
            elif param_type == 'str':
                dpg.add_input_text(tag=tag, default_value=default if default is not None else "",
                                   width=-1)
            elif param_type == 'bool':
                dpg.add_checkbox(tag=tag, default_value=default if default is not None else False)
            else:
                dpg.add_input_text(tag=tag, default_value=str(default) if default is not None else "",
                                   width=-1)


def _get_algo_params(algo_name: str, idx: int) -> dict:
    """Get parameter values for a specific algorithm instance."""
    params = {}
    cat = _state["current_algo_category"]
    algo_params = get_algorithm_params_from_scan(cat, algo_name)

    # Use safe name and index for tag lookup (replace "/" with "-")
    safe_algo_name = algo_name.replace("/", "-")

    for param_name, param_info in algo_params.items():
        tag = f"batch_algo_param_{safe_algo_name}_{idx}_{param_name}"
        if dpg.does_item_exist(tag):
            value = dpg.get_value(tag)
            # Parse value (handles list input for vectorizable params)
            params[param_name] = _parse_param_value(value, param_info.get('type', 'int'))

    return params


def _select_all_algos(sender, app_data):
    """Select all algorithms from current category."""
    cat = _state["current_algo_category"]
    algos = get_algorithm_names(cat)
    _state["selected_algos"] = list(algos)
    _update_algo_params_display()


def _deselect_all_algos(sender, app_data):
    """Clear all selected algorithms."""
    _state["selected_algos"] = []
    _update_algo_params_display()


def _run_clicked(sender, app_data):
    """Run batch experiment."""
    if _state["status"] and _state["status"].running:
        return

    prob_cat = dpg.get_value("batch_prob_cat_combo")
    algo_cat = dpg.get_value("batch_algo_cat_combo")

    n_runs = dpg.get_value("batch_nruns_input")
    max_workers = dpg.get_value("batch_workers_input")

    # Get analysis settings
    table_format = dpg.get_value("batch_tbl_combo")
    figure_format = dpg.get_value("batch_fig_combo")
    statistic_type = dpg.get_value("batch_stat_combo")
    log_scale = dpg.get_value("batch_log_check")
    rank_sum_test = True
    show_pf = dpg.get_value("batch_showpf_check")
    show_nd = dpg.get_value("batch_shownd_check")
    merge_columns = dpg.get_value("batch_mergecols_input")
    # When merge_columns is 0, disable merge_plots
    merge_plots = dpg.get_value("batch_merge_check") and merge_columns > 0
    convergence_k = dpg.get_value("batch_convk_input") or None  # 0 means disabled
    best_so_far = dpg.get_value("batch_bestsofar_check")
    show_std_band = dpg.get_value("batch_stdband_check") if dpg.does_item_exist("batch_stdband_check") else False

    # Get selections from state
    selected_probs = _state["selected_probs"]
    selected_algos = _state["selected_algos"]

    if not selected_probs or not selected_algos:
        show_error_modal("Select at least one problem and one algorithm.")
        return

    data_path = _state["file_manager"].get_data_path_str()
    _state["load_data_save_path"] = None
    save_path = _state["file_manager"].get_results_path_str()
    _state["displayed"] = False

    total_tasks = len(selected_probs) * len(selected_algos) * n_runs
    status = BatchStatus(total_tasks=total_tasks)
    _state["status"] = status

    # Lock run button
    dpg.configure_item("batch_run_btn", enabled=False)
    dpg.bind_item_theme("batch_run_btn", _get_disabled_btn_theme())

    dpg.delete_item("batch_results_area", children_only=True)
    _state["stdout_buffer"] = ""

    # Show progress bar and elapsed time
    dpg.add_text("Experiment running...", parent="batch_results_area", tag="batch_running_text",
                 color=(60, 60, 60))
    dpg.add_progress_bar(tag="batch_progress", parent="batch_results_area",
                         default_value=0.0, width=-1)
    dpg.add_text("Phase: Initializing...", parent="batch_results_area", tag="batch_phase_text",
                 color=(80, 80, 80))
    dpg.add_text("Elapsed: 0s", parent="batch_results_area", tag="batch_elapsed_text",
                 color=(100, 100, 100))

    # Build SETTINGS for metric calculation (always provide so MO data works)
    metric = dpg.get_value("batch_metric_combo") if dpg.does_item_exist("batch_metric_combo") else "IGD"
    settings = {'metric': metric, 'n_ref': 10000}
    if is_multi_objective(prob_cat):
        # Load Pareto front references from suite for MO categories
        problem_names = []
        suite_settings_cache = {}
        for i, (suite, method) in enumerate(selected_probs):
            custom_name = _state["prob_names"].get(i, method)
            problem_names.append(custom_name)
            # Load suite settings once per suite
            if suite not in suite_settings_cache:
                suite_settings_cache[suite] = get_problem_settings(prob_cat, suite) or {}
            suite_st = suite_settings_cache[suite]
            # Copy the method's reference data under the custom_name key
            if method in suite_st:
                settings[custom_name] = suite_st[method]
        settings['problems'] = problem_names

    def _worker():
        capture = io.StringIO()
        status.running = True
        status.start_time = time.time()
        status.current_phase = "running"
        with contextlib.redirect_stdout(capture), contextlib.redirect_stderr(capture):
            try:
                from ddmtolab.Methods.batch_experiment import BatchExperiment
                batch_exp = BatchExperiment(base_path=data_path, clear_folder=True)

                # Add problems with their individual parameters
                for i, (suite, method) in enumerate(selected_probs):
                    # Get problem parameters for this problem
                    prob_params = _state["prob_params"].get(i, {})
                    # Get custom name
                    custom_name = _state["prob_names"].get(i, method)

                    try:
                        creator, _ = get_problem_creator(prob_cat, suite, method, **prob_params)
                    except TypeError:
                        creator, _ = get_problem_creator(prob_cat, suite, method)
                    batch_exp.add_problem(creator, custom_name)

                algo_display_names = []
                for i, algo_name in enumerate(selected_algos):
                    algo_cls = get_algorithm_class(algo_cat, algo_name)
                    # Get algorithm-specific parameters for this instance
                    algo_kwargs = _get_algo_params(algo_name, i)
                    algo_kwargs["disable_tqdm"] = True
                    # Use custom name if set, otherwise use algorithm name
                    custom_name = _state["algo_names"].get(i, algo_name)
                    # Ensure unique names for same algorithms
                    if custom_name == algo_name and selected_algos.count(algo_name) > 1:
                        display_name = f"{algo_name}_{i+1}"
                    else:
                        display_name = custom_name
                    # Sanitize "/" in names for safe file paths
                    safe_display_name = display_name.replace("/", "-")
                    algo_display_names.append(safe_display_name)
                    batch_exp.add_algorithm(algo_cls, safe_display_name, **algo_kwargs)

                batch_exp.run(n_runs=n_runs, max_workers=max_workers)
                _state["stdout_buffer"] = capture.getvalue()

                if status.cancelled:
                    status.error = "Cancelled by user"
                    status.finished = True
                    return

                status.current_phase = "analyzing"

                from ddmtolab.Methods.data_analysis import DataAnalyzer
                analyzer = DataAnalyzer(
                    data_path=data_path,
                    save_path=save_path,
                    settings=settings,
                    algorithm_order=algo_display_names,
                    table_format=table_format,
                    figure_format=figure_format,
                    statistic_type=statistic_type,
                    rank_sum_test=rank_sum_test,
                    log_scale=log_scale,
                    show_pf=show_pf,
                    show_nd=show_nd,
                    merge_plots=merge_plots,
                    merge_columns=merge_columns,
                    best_so_far=best_so_far,
                    clear_results=True,
                    convergence_k=convergence_k,
                    show_std_band=show_std_band,
                )
                status.analysis_result = analyzer.run()

                status.current_phase = "complete"
                status.finished = True
            except Exception as e:
                import traceback
                status.error = f"{str(e)}\n{traceback.format_exc()}"
                status.finished = True
            finally:
                status.running = False
                status.elapsed = time.time() - status.start_time
        _state["stdout_buffer"] = capture.getvalue()

    t = threading.Thread(target=_worker, daemon=True)
    t.start()


def _stop_clicked(sender, app_data):
    """Attempt to signal stop."""
    status = _state.get("status")
    if status and status.running:
        status.cancelled = True
        status.running = False
        status.finished = True
        status.error = "Stopped by user"


def _clean_clicked(sender, app_data):
    """Show cleanup confirmation dialog."""
    def on_backup_and_clean():
        backup_path = _state["backup_manager"].clean_and_backup()
        if backup_path:
            show_info_modal(f"Backup created at:\n{backup_path}\n\nData cleaned.")
        else:
            show_info_modal("No data to backup. Folders cleaned.")
        dpg.delete_item("batch_results_area", children_only=True)
        dpg.add_text("Data cleaned. Ready for new experiment.", parent="batch_results_area",
                     color=(100, 100, 100))

    def on_clean_only():
        _state["backup_manager"].clean_without_backup()
        show_info_modal("Data cleaned without backup.")
        dpg.delete_item("batch_results_area", children_only=True)
        dpg.add_text("Data cleaned. Ready for new experiment.", parent="batch_results_area",
                     color=(100, 100, 100))

    show_confirm_dialog(
        "Do you want to backup data before cleaning?",
        on_yes=on_backup_and_clean,
        on_no=on_clean_only,
        title="Clean Data",
        yes_label="Backup & Clean",
        no_label="Clean Only",
        cancel_label="Cancel"
    )


def _save_config(sender, app_data):
    """Save current experiment configuration to YAML file.

    Format is compatible with BatchExperiment.from_config().
    """
    from datetime import datetime
    from utils.problem_scanner import get_problem_class

    # Gather current configuration
    prob_cat = dpg.get_value("batch_prob_cat_combo")
    algo_cat = dpg.get_value("batch_algo_cat_combo")
    selected_probs = _state["selected_probs"]
    selected_algos = _state["selected_algos"]

    if not selected_probs or not selected_algos:
        show_error_modal("Please select at least one problem and one algorithm.")
        return

    # Build problems list in BatchExperiment format
    problems_config = []
    for i, (suite, method) in enumerate(selected_probs):
        # Get problem module info via scanner API
        prob_module_path = get_problem_module_path(prob_cat, suite) or ""
        prob_cls = get_problem_class(prob_cat, suite)
        prob_class_name = prob_cls.__name__ if prob_cls else suite
        # Get custom name and parameters
        custom_name = _state["prob_names"].get(i, method)
        prob_params = _state["prob_params"].get(i, {})

        problems_config.append({
            'name': custom_name,
            'class': prob_class_name,
            'creator_name': method,
            'module': prob_module_path,
            'suite': suite,
            'params': prob_params.copy()
        })

    # Build algorithms list in BatchExperiment format
    algorithms_config = []
    for i, algo_name in enumerate(selected_algos):
        algo_params = _get_algo_params(algo_name, i)
        custom_name = _state["algo_names"].get(i, algo_name)

        # Ensure unique names for same algorithms
        if custom_name == algo_name and selected_algos.count(algo_name) > 1:
            display_name = f"{algo_name}_{i+1}"
        else:
            display_name = custom_name

        # Get algorithm module info via scanner API
        algo_info = get_discovered_algorithm_module_info(algo_cat, algo_name)
        algo_module_path = algo_info[0] if algo_info else ""
        algo_class_name = algo_info[1] if algo_info else algo_name

        algorithms_config.append({
            'name': display_name,
            'class': algo_class_name,
            'module': algo_module_path,
            'parameters': algo_params
        })

    # Build config in BatchExperiment format
    base_path = _state["file_manager"].get_data_path_str()
    config = {
        'created_at': datetime.now().isoformat(),
        'base_path': base_path,
        'clear_folder': False,
        'problems': problems_config,
        'algorithms': algorithms_config,
        'run_settings': {
            'n_runs': dpg.get_value("batch_nruns_input"),
            'max_workers': dpg.get_value("batch_workers_input"),
        },
        # UI-specific settings (not used by BatchExperiment but preserved for UI reload)
        'ui_settings': {
            'problem_category': prob_cat,
            'algorithm_category': algo_cat,
            'metric': dpg.get_value("batch_metric_combo") if dpg.does_item_exist("batch_metric_combo") else "IGD",
        },
        'analysis_settings': {
            'table_format': dpg.get_value("batch_tbl_combo"),
            'figure_format': dpg.get_value("batch_fig_combo"),
            'statistic_type': dpg.get_value("batch_stat_combo"),
            'log_scale': dpg.get_value("batch_log_check"),
            'show_pf': dpg.get_value("batch_showpf_check"),
            'show_nd': dpg.get_value("batch_shownd_check"),
            'merge_plots': dpg.get_value("batch_merge_check"),
            'merge_columns': dpg.get_value("batch_mergecols_input"),
            'convergence_k': dpg.get_value("batch_convk_input"),
            'best_so_far': dpg.get_value("batch_bestsofar_check"),
            'show_std_band': dpg.get_value("batch_stdband_check") if dpg.does_item_exist("batch_stdband_check") else False,
        }
    }

    # Save to file with custom formatting (same as BatchExperiment.save_config)
    config_path = Path(_state["file_manager"].base_path) / "experiment_config.yaml"
    try:
        with open(config_path, 'w', encoding='utf-8') as f:
            # Write basic info
            f.write(f"created_at: {config['created_at']}\n")
            f.write(f"base_path: {config['base_path']}\n")
            f.write(f"clear_folder: {config['clear_folder']}\n\n")

            # Write problems with blank lines between each
            f.write("problems:\n")
            for i, prob in enumerate(config['problems']):
                if i > 0:
                    f.write("\n")
                f.write(f"  - name: {prob['name']}\n")
                f.write(f"    class: {prob['class']}\n")
                f.write(f"    creator_name: {prob['creator_name']}\n")
                f.write(f"    module: {prob['module']}\n")
                f.write(f"    suite: {prob['suite']}\n")
                f.write(f"    params: {prob['params']}\n")

            # Write algorithms with blank lines between each
            f.write("\nalgorithms:\n")
            for i, algo in enumerate(config['algorithms']):
                if i > 0:
                    f.write("\n")
                f.write(f"  - name: {algo['name']}\n")
                f.write(f"    class: {algo['class']}\n")
                f.write(f"    module: {algo['module']}\n")
                f.write(f"    parameters:\n")
                for key, value in algo['parameters'].items():
                    f.write(f"      {key}: {value}\n")

            # Write run settings
            f.write("\nrun_settings:\n")
            f.write(f"  n_runs: {config['run_settings']['n_runs']}\n")
            f.write(f"  max_workers: {config['run_settings']['max_workers']}\n")

            # Write UI settings
            f.write("\nui_settings:\n")
            for key, value in config['ui_settings'].items():
                f.write(f"  {key}: {value}\n")

            # Write analysis settings
            f.write("\nanalysis_settings:\n")
            for key, value in config['analysis_settings'].items():
                f.write(f"  {key}: {value}\n")

        show_info_modal(f"Configuration saved to:\n{config_path}")
    except Exception as e:
        show_error_modal(f"Failed to save config: {e}")


def _load_config(sender, app_data):
    """Open native file dialog to select and load experiment configuration from YAML file."""
    import tkinter as tk
    from tkinter import filedialog

    # Create a hidden tkinter root window
    root = tk.Tk()
    root.withdraw()  # Hide the root window
    root.attributes('-topmost', True)  # Bring dialog to front

    # Default path is tests folder
    default_path = str(_state["file_manager"].base_path)

    # Open native file dialog
    file_path = filedialog.askopenfilename(
        title="Select Configuration File",
        initialdir=default_path,
        filetypes=[
            ("YAML files", "*.yaml *.yml"),
            ("All files", "*.*")
        ]
    )

    root.destroy()  # Clean up tkinter

    # Load the selected file
    if file_path:
        _load_config_from_file(file_path)


def _load_config_from_file(config_path: str):
    """Load experiment configuration from specified YAML file.

    Compatible with BatchExperiment.save_config() format.
    """
    try:
        import yaml
    except ImportError:
        show_error_modal("Please install PyYAML: pip install pyyaml")
        return

    config_path = Path(config_path)
    if not config_path.exists():
        show_error_modal(f"Config file not found:\n{config_path}")
        return

    try:
        with open(config_path, 'r', encoding='utf-8') as f:
            config = yaml.safe_load(f)
    except Exception as e:
        show_error_modal(f"Failed to load config: {e}")
        return

    # Helper function: find category and suite from module path
    def find_problem_category_and_suite(module_path: str):
        from utils.registry import get_all_categories, get_problem_suites, get_problem_module_path
        for cat in get_all_categories():
            for suite_name in get_problem_suites(cat):
                if get_problem_module_path(cat, suite_name) == module_path:
                    return cat, suite_name
        return None, None

    # Helper function: find category and display name from module path and class name
    def find_algorithm_category_and_name(module_path: str, class_name: str):
        from utils.registry import get_all_categories, get_algorithm_names
        for cat in get_all_categories():
            for name in get_algorithm_names(cat):
                info = get_discovered_algorithm_module_info(cat, name)
                if info and info[0] == module_path and info[1] == class_name:
                    return cat, name
        return None, None

    # Check if config has ui_settings (UI-generated config)
    ui_settings = config.get('ui_settings', {})

    # Determine problem and algorithm categories
    problems = config.get('problems', [])
    algorithms = config.get('algorithms', [])

    if not problems:
        show_error_modal("No problems found in config.")
        return
    if not algorithms:
        show_error_modal("No algorithms found in config.")
        return

    # Get problem category from ui_settings or infer from module path
    if ui_settings:
        prob_cat = ui_settings.get('problem_category', 'STSO')
        algo_cat = ui_settings.get('algorithm_category', 'STSO')
    else:
        prob_cat, _ = find_problem_category_and_suite(problems[0].get('module', ''))
        if not prob_cat:
            show_error_modal("Could not determine problem category from config.")
            return
        algo_cat, _ = find_algorithm_category_and_name(
            algorithms[0].get('module', ''), algorithms[0].get('class', ''))
        if not algo_cat:
            show_error_modal("Could not determine algorithm category from config.")
            return

    # Apply problem category
    dpg.set_value("batch_prob_cat_combo", prob_cat)
    _on_prob_category_change(None, prob_cat)

    # Clear and populate selected problems
    _state["selected_probs"] = []
    _state["prob_names"] = {}
    _state["prob_params"] = {}

    for i, prob_cfg in enumerate(problems):
        # Get suite from config or infer from module path
        suite = prob_cfg.get('suite', '')
        if not suite:
            _, suite = find_problem_category_and_suite(prob_cfg.get('module', ''))

        if not suite:
            continue

        method = prob_cfg.get('creator_name', '')
        if not method:
            continue

        # Add to selected problems
        _state["selected_probs"].append((suite, method))

        # Set custom name if different from default
        custom_name = prob_cfg.get('name', method)
        if custom_name != method:
            _state["prob_names"][i] = custom_name

        # Set parameters
        params = prob_cfg.get('params', {})
        if params:
            _state["prob_params"][i] = params.copy()

    # Update suite combo to first problem's suite (for the problem list display)
    if _state["selected_probs"]:
        first_suite = _state["selected_probs"][0][0]
        dpg.set_value("batch_suite_combo", first_suite)
        _on_suite_change(None, first_suite)

    # Update problem params display
    _update_prob_params_display()

    # Apply metric from ui_settings
    if ui_settings.get('metric') and dpg.does_item_exist("batch_metric_combo"):
        dpg.set_value("batch_metric_combo", ui_settings['metric'])

    # Apply algorithm category
    dpg.set_value("batch_algo_cat_combo", algo_cat)
    _state["current_algo_category"] = algo_cat
    _update_algo_list_display()

    # Clear and set selected algorithms
    _state["selected_algos"] = []
    _state["algo_names"] = {}

    algorithms = config.get('algorithms', [])
    for i, algo_cfg in enumerate(algorithms):
        # Find algorithm display name from module and class
        algo_module = algo_cfg.get('module', '')
        algo_class = algo_cfg.get('class', '')
        _, algo_display_name = find_algorithm_category_and_name(algo_module, algo_class)

        if algo_display_name:
            _state["selected_algos"].append(algo_display_name)
            # Set custom name if different from display name
            custom_name = algo_cfg.get('name', algo_display_name)
            if custom_name != algo_display_name:
                _state["algo_names"][i] = custom_name

    _update_algo_params_display()

    # Apply algorithm parameters after display is updated
    for i, algo_cfg in enumerate(algorithms):
        algo_module = algo_cfg.get('module', '')
        algo_class = algo_cfg.get('class', '')
        _, algo_display_name = find_algorithm_category_and_name(algo_module, algo_class)

        if not algo_display_name:
            continue

        params = algo_cfg.get('parameters', {})
        safe_algo_name = algo_display_name.replace("/", "-")
        for param_name, value in params.items():
            tag = f"batch_algo_param_{safe_algo_name}_{i}_{param_name}"
            if dpg.does_item_exist(tag):
                if isinstance(value, list):
                    dpg.set_value(tag, str(value))
                else:
                    dpg.set_value(tag, value)

    # Apply run settings
    run_config = config.get('run_settings', {})
    if 'n_runs' in run_config:
        dpg.set_value("batch_nruns_input", run_config['n_runs'])
    if 'max_workers' in run_config:
        dpg.set_value("batch_workers_input", run_config['max_workers'])

    # Apply analysis settings
    analysis_config = config.get('analysis_settings', {})
    setting_map = {
        'table_format': 'batch_tbl_combo',
        'figure_format': 'batch_fig_combo',
        'statistic_type': 'batch_stat_combo',
        'log_scale': 'batch_log_check',
        'show_pf': 'batch_showpf_check',
        'show_nd': 'batch_shownd_check',
        'merge_plots': 'batch_merge_check',
        'merge_columns': 'batch_mergecols_input',
        'convergence_k': 'batch_convk_input',
        'best_so_far': 'batch_bestsofar_check',
        'show_std_band': 'batch_stdband_check',
    }
    for key, tag in setting_map.items():
        if key in analysis_config and dpg.does_item_exist(tag):
            dpg.set_value(tag, analysis_config[key])

    show_info_modal("Configuration loaded successfully!")


def _display_analysis_results(save_path: str, merge_columns: int = 0):
    """Display analysis outputs."""
    results_dir = Path(save_path)
    if not results_dir.exists():
        dpg.add_text("Results folder not found.", parent="batch_results_area", color=(180, 80, 80))
        return

    # Image layout: merged convergence → 1 per row full width; otherwise 3 per row
    viewport_w = dpg.get_viewport_width()
    available_width = int(max(400, viewport_w - 250 - 280 - 50) * 0.9)
    IMG_WIDTH = (available_width - 40) // 3  # 3 images + spacing per row
    per_row = 3

    if merge_columns > 0:
        conv_img_width = available_width - 20
        conv_per_row = 1
        nd_img_width = available_width - 20
        nd_per_row = 1
    else:
        conv_img_width = IMG_WIDTH
        conv_per_row = per_row
        nd_img_width = IMG_WIDTH
        nd_per_row = per_row

    # Open folder button at top right (with dark text theme)
    def _open_folder():
        import subprocess
        import sys
        if sys.platform == 'win32':
            os.startfile(str(results_dir.resolve()))
        elif sys.platform == 'darwin':
            subprocess.run(['open', str(results_dir.resolve())])
        else:
            subprocess.run(['xdg-open', str(results_dir.resolve())])

    with dpg.group(horizontal=True, parent="batch_results_area"):
        dpg.add_spacer(width=-1)
        btn = dpg.add_button(label="Open Results Folder", callback=lambda: _open_folder(), width=150)
        # Apply dark text theme to button
        with dpg.theme() as btn_theme:
            with dpg.theme_component(dpg.mvButton):
                dpg.add_theme_color(dpg.mvThemeCol_Text, (40, 40, 40))
                dpg.add_theme_color(dpg.mvThemeCol_Button, (200, 200, 200))
                dpg.add_theme_color(dpg.mvThemeCol_ButtonHovered, (180, 180, 180))
                dpg.add_theme_color(dpg.mvThemeCol_ButtonActive, (160, 160, 160))
        dpg.bind_item_theme(btn, btn_theme)

    dpg.add_spacer(height=5, parent="batch_results_area")

    # Excel/LaTeX tables
    dpg.add_text("Statistical Tables", parent="batch_results_area", color=COLOR_TITLE)
    dpg.add_separator(parent="batch_results_area")

    # Create a theme for table - header row needs white text on dark background
    with dpg.theme() as table_theme:
        with dpg.theme_component(dpg.mvTable):
            dpg.add_theme_color(dpg.mvThemeCol_TableHeaderBg, (70, 70, 70))
            dpg.add_theme_color(dpg.mvThemeCol_Text, (255, 255, 255))  # White text for header
        with dpg.theme_component(dpg.mvTableRow):
            dpg.add_theme_color(dpg.mvThemeCol_Text, (40, 40, 40))  # Dark text for rows

    excel_files = sorted(results_dir.glob("*.xlsx")) + sorted(results_dir.glob("*.xls"))
    if excel_files:
        for f in excel_files:
            dpg.add_text(f.stem, parent="batch_results_area", color=COLOR_SECTION)
            try:
                from openpyxl import load_workbook
                import pandas as pd

                # Load workbook to get formatting info
                wb = load_workbook(f)
                ws = wb.active

                # Build a set of bold cells (row, col) - 1-indexed
                bold_cells = set()
                for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=1):  # Skip header
                    for col_idx, cell in enumerate(row, start=0):
                        if cell.font and cell.font.bold:
                            bold_cells.add((row_idx, col_idx))

                # Read data with pandas
                df = pd.read_excel(f)
                # Filter out unnamed or empty columns
                df = df.loc[:, ~df.columns.str.contains('^Unnamed')]
                df = df.dropna(axis=1, how='all')
                cols = list(df.columns)

                tbl = dpg.add_table(parent="batch_results_area", header_row=True,
                               borders_innerH=True, borders_outerH=True,
                               borders_innerV=True, borders_outerV=True,
                               policy=dpg.mvTable_SizingStretchSame,
                               no_host_extendX=True, precise_widths=True)
                dpg.bind_item_theme(tbl, table_theme)
                for i, col in enumerate(cols):
                    if i == 0:
                        col_width = 150  # First column
                    elif i == 1:
                        col_width = 50   # Second column
                    else:
                        col_width = 180  # Algorithm columns
                    dpg.add_table_column(label=str(col), parent=tbl, width_fixed=True, init_width_or_weight=col_width)

                for row_idx, row in df.iterrows():
                    with dpg.table_row(parent=tbl):
                        for col_idx, col in enumerate(cols):
                            cell_text = str(row[col])
                            # Check if this cell is bold in Excel (row_idx is 0-based, +1 for Excel)
                            is_bold = (row_idx + 1, col_idx) in bold_cells
                            if is_bold:
                                # Highlight bold cells with blue color (same as title)
                                dpg.add_text(cell_text, color=COLOR_TITLE)
                            else:
                                dpg.add_text(cell_text)
            except Exception as e:
                dpg.add_text(f"Could not read {f.name}: {e}", parent="batch_results_area",
                             color=(200, 60, 60))
    else:
        tex_files = sorted(results_dir.glob("*.tex")) + sorted(results_dir.glob("*.txt"))
        for f in tex_files:
            try:
                content = f.read_text(encoding="utf-8")
                with dpg.group(horizontal=True, parent="batch_results_area"):
                    dpg.add_text(f.stem, color=COLOR_SECTION)
                    copy_btn = dpg.add_button(
                        label="Copy",
                        callback=lambda s, a, u: copy_text_to_clipboard(u),
                        user_data=content, width=50,
                    )
                    with dpg.theme() as copy_btn_theme:
                        with dpg.theme_component(dpg.mvButton):
                            dpg.add_theme_color(dpg.mvThemeCol_Text, (40, 40, 40))
                            dpg.add_theme_color(dpg.mvThemeCol_Button, (200, 200, 200))
                            dpg.add_theme_color(dpg.mvThemeCol_ButtonHovered, (180, 180, 180))
                            dpg.add_theme_color(dpg.mvThemeCol_ButtonActive, (160, 160, 160))
                    dpg.bind_item_theme(copy_btn, copy_btn_theme)
                # Calculate height based on line count (approx 18px per line)
                line_count = content.count('\n') + 1
                text_height = line_count * 18 + 10
                dpg.add_input_text(default_value=content, multiline=True, readonly=True,
                                   height=text_height, width=-1, parent="batch_results_area")
            except Exception:
                pass

    def _add_image_with_menu(img_path, parent=None, max_width=280):
        """Add an image with right-click menu."""
        tex = load_image_to_texture(str(img_path))
        if tex:
            tag, w, h = tex
            if w > max_width:
                h = int(h * max_width / w)
                w = max_width
            if parent is not None:
                img_widget = dpg.add_image(tag, width=w, height=h, parent=parent)
            else:
                img_widget = dpg.add_image(tag, width=w, height=h)
            with dpg.popup(img_widget, mousebutton=dpg.mvMouseButton_Right):
                dpg.add_menu_item(label="Copy Image",
                                  callback=_copy_image_to_clipboard, user_data=str(img_path))
                dpg.add_menu_item(label="Open File Location",
                                  callback=_open_file_location, user_data=str(img_path))
            return True
        return False

    # Collect images by category from root directory
    all_root_images = sorted(results_dir.glob("*.png"))
    runtime_images = [f for f in all_root_images if "runtime" in f.name.lower()]
    # All other root images are convergence curves
    convergence_images = [f for f in all_root_images if f not in runtime_images]

    # ND Solutions from subdirectory
    nd_folder = results_dir / "ND_Solutions"
    nd_images = sorted(nd_folder.glob("*.png")) if nd_folder.exists() else []

    # 1. Display convergence images (merged: 1 per row full width; non-merged: adaptive)
    if convergence_images:
        dpg.add_spacer(height=10, parent="batch_results_area")
        dpg.add_text("Convergence Plots", parent="batch_results_area", color=COLOR_TITLE)
        dpg.add_separator(parent="batch_results_area")
        for i in range(0, len(convergence_images), conv_per_row):
            row_files = convergence_images[i:i+conv_per_row]
            with dpg.group(horizontal=True, parent="batch_results_area"):
                for img_path in row_files:
                    _add_image_with_menu(img_path, max_width=conv_img_width)
                    dpg.add_spacer(width=10)
            dpg.add_spacer(height=8, parent="batch_results_area")

    # 2. Display ND Solutions (adaptive per row)
    if nd_images:
        dpg.add_spacer(height=10, parent="batch_results_area")
        dpg.add_text("Non-Dominated Solutions", parent="batch_results_area", color=COLOR_TITLE)
        dpg.add_separator(parent="batch_results_area")
        for i in range(0, len(nd_images), nd_per_row):
            row_files = nd_images[i:i+nd_per_row]
            with dpg.group(horizontal=True, parent="batch_results_area"):
                for img_path in row_files:
                    _add_image_with_menu(img_path, max_width=nd_img_width)
                    dpg.add_spacer(width=10)
            dpg.add_spacer(height=8, parent="batch_results_area")

    # 3. Display runtime images (adaptive per row)
    if runtime_images:
        dpg.add_spacer(height=10, parent="batch_results_area")
        dpg.add_text("Runtime Comparison", parent="batch_results_area", color=COLOR_TITLE)
        dpg.add_separator(parent="batch_results_area")
        for i in range(0, len(runtime_images), per_row):
            row_files = runtime_images[i:i+per_row]
            with dpg.group(horizontal=True, parent="batch_results_area"):
                for img_path in row_files:
                    _add_image_with_menu(img_path, max_width=IMG_WIDTH)
                    dpg.add_spacer(width=10)
            dpg.add_spacer(height=8, parent="batch_results_area")

    if not convergence_images and not runtime_images and not nd_images:
        other = sorted(results_dir.glob("*.pdf")) + sorted(results_dir.glob("*.svg"))
        if other:
            dpg.add_text(f"Found {len(other)} plot files (PDF/SVG). Check {save_path} folder.",
                         parent="batch_results_area", color=(100, 100, 100))


def _load_data_clicked(sender, app_data):
    """Open folder dialog to load existing experiment data and run analysis."""
    import tkinter as tk
    from tkinter import filedialog

    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)

    # Open parent folder (tests/) so that Data folder is visible and selectable
    default_path = str(Path(_state["file_manager"].get_data_path_str()).parent)

    folder_path = filedialog.askdirectory(
        title="Select Data Folder",
        initialdir=default_path,
    )

    root.destroy()

    if not folder_path:
        return

    # Check if folder has .pkl files
    data_dir = Path(folder_path)
    pkl_files = list(data_dir.rglob("*.pkl"))
    if not pkl_files:
        show_error_modal("No .pkl data files found in selected folder.")
        return

    # Set save path as sibling Results folder
    save_path = str(data_dir.parent / "Results")
    _state["load_data_save_path"] = save_path

    # Get analysis settings from UI
    table_format = dpg.get_value("batch_tbl_combo")
    figure_format = dpg.get_value("batch_fig_combo")
    statistic_type = dpg.get_value("batch_stat_combo")
    log_scale = dpg.get_value("batch_log_check")
    show_pf = dpg.get_value("batch_showpf_check")
    show_nd = dpg.get_value("batch_shownd_check")
    merge_columns = dpg.get_value("batch_mergecols_input")
    merge_plots = dpg.get_value("batch_merge_check") and merge_columns > 0
    convergence_k = dpg.get_value("batch_convk_input") or None  # 0 means disabled
    best_so_far = dpg.get_value("batch_bestsofar_check")
    show_std_band = dpg.get_value("batch_stdband_check") if dpg.does_item_exist("batch_stdband_check") else False

    _state["displayed"] = False

    status = BatchStatus(total_tasks=0)
    _state["status"] = status

    # Lock run button
    dpg.configure_item("batch_run_btn", enabled=False)
    dpg.bind_item_theme("batch_run_btn", _get_disabled_btn_theme())

    dpg.delete_item("batch_results_area", children_only=True)
    dpg.add_text("Analyzing loaded data...", parent="batch_results_area",
                 tag="batch_running_text", color=(60, 60, 60))
    dpg.add_progress_bar(tag="batch_progress", parent="batch_results_area",
                         default_value=0.5, width=-1)
    dpg.add_text("Phase: Analyzing...", parent="batch_results_area",
                 tag="batch_phase_text", color=(80, 80, 80))
    dpg.add_text("", parent="batch_results_area",
                 tag="batch_elapsed_text", color=(100, 100, 100))

    # Build settings for metric calculation with reference data from all MO suites
    metric = dpg.get_value("batch_metric_combo") if dpg.does_item_exist("batch_metric_combo") else "IGD"
    settings = {'metric': metric, 'n_ref': 10000}
    # Load Pareto front references from all MO problem suites so loaded data can use them
    for mo_cat in ['STMO', 'MTMO']:
        for suite in get_problem_suites(mo_cat):
            suite_settings = get_problem_settings(mo_cat, suite)
            if suite_settings:
                for key, value in suite_settings.items():
                    if key not in ('metric', 'n_ref') and key not in settings:
                        settings[key] = value

    def _worker():
        status.running = True
        status.start_time = time.time()
        status.current_phase = "analyzing"
        try:
            from ddmtolab.Methods.data_analysis import DataAnalyzer
            analyzer = DataAnalyzer(
                data_path=folder_path,
                save_path=save_path,
                settings=settings,
                table_format=table_format,
                figure_format=figure_format,
                statistic_type=statistic_type,
                rank_sum_test=True,
                log_scale=log_scale,
                show_pf=show_pf,
                show_nd=show_nd,
                merge_plots=merge_plots,
                merge_columns=merge_columns,
                best_so_far=best_so_far,
                clear_results=True,
                convergence_k=convergence_k,
                show_std_band=show_std_band,
            )
            status.analysis_result = analyzer.run()
            status.current_phase = "complete"
            status.finished = True
        except Exception as e:
            import traceback
            status.error = f"{str(e)}\n{traceback.format_exc()}"
            status.finished = True
        finally:
            status.running = False
            status.elapsed = time.time() - status.start_time

    t = threading.Thread(target=_worker, daemon=True)
    t.start()


def create(parent, base_path: str = "./tests"):
    """Create the Batch Mode UI."""
    _init_managers(base_path)

    with dpg.group(horizontal=True, parent=parent):
        # Left panel - Problem Selection (Column 1)
        with dpg.child_window(width=250, tag="batch_problem_panel"):
            dpg.add_text("Problem Selection", color=COLOR_TITLE)
            dpg.add_separator()

            dpg.add_text("Category", color=COLOR_SUCCESS)
            dpg.add_combo(CATEGORIES, default_value="STSO", tag="batch_prob_cat_combo",
                          callback=_on_prob_category_change, width=-1)

            dpg.add_spacer(height=5)
            dpg.add_text("Suite", color=COLOR_SUCCESS)
            initial_suites = get_problem_suites("STSO")
            dpg.add_combo(initial_suites,
                          default_value=initial_suites[0] if initial_suites else "",
                          tag="batch_suite_combo", callback=_on_suite_change, width=-1)

            dpg.add_spacer(height=5)
            with dpg.child_window(tag="batch_prob_list_container", height=150, border=True):
                # Will be populated by _update_prob_list_display
                pass

            # Selected Problems with parameters
            dpg.add_spacer(height=10)
            dpg.add_text("Selected Problems", color=COLOR_TITLE)
            dpg.add_separator()
            with dpg.child_window(tag="batch_prob_params_container", height=-1, border=False):
                dpg.add_text("Click problems above to add", color=(150, 150, 150))

        # Middle panel - Algorithm Selection (Column 2)
        with dpg.child_window(width=280, tag="batch_algo_panel"):
            dpg.add_text("Algorithm Selection", color=COLOR_TITLE)
            dpg.add_separator()

            dpg.add_text("Category", color=COLOR_SUCCESS)
            dpg.add_combo(ALGO_CATEGORIES, default_value="STSO", tag="batch_algo_cat_combo",
                          callback=_on_algo_category_change, width=-1)

            dpg.add_spacer(height=5)
            with dpg.child_window(tag="batch_algo_list_container", height=215, border=True):
                # Will be populated by _update_algo_list_display
                pass

            # Algorithm Parameters (dynamic) - includes selected algorithms with delete buttons
            dpg.add_spacer(height=10)
            dpg.add_text("Selected Algorithms", color=COLOR_TITLE)
            dpg.add_separator()
            with dpg.child_window(tag="batch_algo_params_container", height=-1, border=False):
                dpg.add_text("Click algorithms above to add", color=(150, 150, 150))

        # Right panel - Settings and Results
        with dpg.child_window(tag="batch_right_panel"):
            # Run Options and Buttons at the top
            with dpg.theme() as _input_align_theme:
                with dpg.theme_component(dpg.mvAll):
                    dpg.add_theme_style(dpg.mvStyleVar_FramePadding, 8, 6)

            with dpg.group(horizontal=True):
                clean_btn = dpg.add_button(label="Clean Data", callback=_clean_clicked, width=95)
                dpg.add_spacer(width=4)
                load_btn = dpg.add_button(label="Load Data", callback=_load_data_clicked, width=90)
                dpg.add_spacer(width=4)
                savecfg_btn = dpg.add_button(label="Save Config", callback=_save_config, width=100)
                dpg.add_spacer(width=4)
                loadcfg_btn = dpg.add_button(label="Load Config", callback=_load_config, width=100)
                dpg.add_spacer(width=8)
                dpg.add_text("Runs")
                nruns_input = dpg.add_input_int(default_value=3, min_value=1, max_value=100,
                                  tag="batch_nruns_input", width=50, step=0)
                dpg.add_spacer(width=4)
                dpg.add_text("Workers")
                workers_input = dpg.add_input_int(default_value=6, min_value=1, max_value=32,
                                  tag="batch_workers_input", width=50, step=0)
                dpg.add_spacer(width=4)
                run_btn = dpg.add_button(label="Run", tag="batch_run_btn",
                                         callback=_run_clicked, width=100)
                dpg.add_spacer(width=4)
                stop_btn = dpg.add_button(label="Stop", tag="batch_stop_btn",
                                          callback=_stop_clicked, width=100)
            dpg.bind_item_theme(nruns_input, _input_align_theme)
            dpg.bind_item_theme(workers_input, _input_align_theme)

            # Secondary buttons (Clean Data, Load Data, Save Config, Load Config)
            with dpg.theme() as secondary_theme:
                with dpg.theme_component(dpg.mvButton):
                    dpg.add_theme_color(dpg.mvThemeCol_Button, (55, 75, 100))
                    dpg.add_theme_color(dpg.mvThemeCol_ButtonHovered, (70, 95, 125))
                    dpg.add_theme_color(dpg.mvThemeCol_ButtonActive, (85, 115, 150))
                    dpg.add_theme_style(dpg.mvStyleVar_FramePadding, 8, 6)
                    dpg.add_theme_style(dpg.mvStyleVar_FrameRounding, 4)
                    dpg.add_theme_style(dpg.mvStyleVar_FrameBorderSize, 1)
                    dpg.add_theme_color(dpg.mvThemeCol_Border, (90, 110, 140))
            dpg.bind_item_theme(clean_btn, secondary_theme)
            dpg.bind_item_theme(load_btn, secondary_theme)
            dpg.bind_item_theme(savecfg_btn, secondary_theme)
            dpg.bind_item_theme(loadcfg_btn, secondary_theme)

            # Primary button (Run)
            with dpg.theme() as run_theme:
                with dpg.theme_component(dpg.mvButton):
                    dpg.add_theme_color(dpg.mvThemeCol_Button, (40, 140, 70))
                    dpg.add_theme_color(dpg.mvThemeCol_ButtonHovered, (55, 165, 90))
                    dpg.add_theme_color(dpg.mvThemeCol_ButtonActive, (70, 185, 110))
                    dpg.add_theme_style(dpg.mvStyleVar_FramePadding, 8, 6)
                    dpg.add_theme_style(dpg.mvStyleVar_FrameRounding, 4)
                    dpg.add_theme_style(dpg.mvStyleVar_FrameBorderSize, 1)
                    dpg.add_theme_color(dpg.mvThemeCol_Border, (60, 170, 90))
            dpg.bind_item_theme(run_btn, run_theme)
            _state["run_theme"] = run_theme

            # Danger button (Stop)
            with dpg.theme() as stop_theme:
                with dpg.theme_component(dpg.mvButton):
                    dpg.add_theme_color(dpg.mvThemeCol_Button, (170, 50, 50))
                    dpg.add_theme_color(dpg.mvThemeCol_ButtonHovered, (195, 70, 70))
                    dpg.add_theme_color(dpg.mvThemeCol_ButtonActive, (215, 95, 95))
                    dpg.add_theme_style(dpg.mvStyleVar_FramePadding, 8, 6)
                    dpg.add_theme_style(dpg.mvStyleVar_FrameRounding, 4)
                    dpg.add_theme_style(dpg.mvStyleVar_FrameBorderSize, 1)
                    dpg.add_theme_color(dpg.mvThemeCol_Border, (200, 70, 70))
            dpg.bind_item_theme(stop_btn, stop_theme)

            with dpg.theme() as delete_theme:
                with dpg.theme_component(dpg.mvButton):
                    dpg.add_theme_color(dpg.mvThemeCol_Button, (140, 50, 50))
                    dpg.add_theme_color(dpg.mvThemeCol_ButtonHovered, (170, 70, 70))
                    dpg.add_theme_color(dpg.mvThemeCol_ButtonActive, (200, 90, 90))
            _state["delete_theme"] = delete_theme

            # Apply bold font to Run and Stop buttons
            from main import get_fonts
            _bold = get_fonts().get("bold")
            if _bold:
                dpg.bind_item_font(run_btn, _bold)
                dpg.bind_item_font(stop_btn, _bold)

            dpg.add_spacer(height=2)
            dpg.add_separator()

            # Analysis Settings
            dpg.add_text("Analysis Settings", color=COLOR_TITLE)
            dpg.add_separator()
            dpg.add_spacer(height=3)

            # Analysis settings table: 2 rows, 6 columns, vertically aligned
            with dpg.theme() as _tbl_pad_theme:
                with dpg.theme_component(dpg.mvTable):
                    dpg.add_theme_style(dpg.mvStyleVar_CellPadding, 12, 4)
            with dpg.table(header_row=False, borders_innerH=False, borders_outerH=False,
                           borders_innerV=False, borders_outerV=False,
                           policy=dpg.mvTable_SizingFixedFit, pad_outerX=False) as settings_tbl:
                dpg.bind_item_theme(settings_tbl, _tbl_pad_theme)
                for _ in range(6):
                    dpg.add_table_column()

                # Row 1: Checkboxes (label before checkbox)
                with dpg.table_row():
                    with dpg.group(horizontal=True):
                        dpg.add_text("Log Scale")
                        dpg.add_checkbox(tag="batch_log_check", default_value=True)
                    with dpg.group(horizontal=True):
                        dpg.add_text("Best-So-Far")
                        dpg.add_checkbox(tag="batch_bestsofar_check", default_value=True)
                    with dpg.group(horizontal=True):
                        dpg.add_text("Show ND")
                        dpg.add_checkbox(tag="batch_shownd_check", default_value=True)
                    with dpg.group(horizontal=True):
                        dpg.add_text("Show PF")
                        dpg.add_checkbox(tag="batch_showpf_check", default_value=True)
                    with dpg.group(horizontal=True):
                        dpg.add_text("Std Band")
                        dpg.add_checkbox(tag="batch_stdband_check", default_value=False)
                    with dpg.group(horizontal=True):
                        dpg.add_text("Merge Plots")
                        dpg.add_checkbox(tag="batch_merge_check", default_value=False)

                # Row 2: Combos and inputs
                with dpg.table_row():
                    with dpg.group(horizontal=True):
                        dpg.add_text("Metric")
                        dpg.add_combo(METRICS, default_value="IGD",
                                      tag="batch_metric_combo", width=80)
                    with dpg.group(horizontal=True):
                        dpg.add_text("Table")
                        dpg.add_combo(TABLE_FORMATS, default_value="excel",
                                      tag="batch_tbl_combo", width=70)
                    with dpg.group(horizontal=True):
                        dpg.add_text("Statistic")
                        dpg.add_combo(STATISTIC_TYPES, default_value="mean",
                                      tag="batch_stat_combo", width=70)
                    with dpg.group(horizontal=True):
                        dpg.add_text("Figure")
                        dpg.add_combo(FIGURE_FORMATS, default_value="png",
                                      tag="batch_fig_combo", width=70)
                    with dpg.group(horizontal=True):
                        dpg.add_text("Conv Samples")
                        dpg.add_input_int(default_value=0, min_value=0, max_value=100,
                                          tag="batch_convk_input", width=50, step=0)
                    with dpg.group(horizontal=True):
                        dpg.add_text("Merge Cols")
                        dpg.add_input_int(default_value=3, min_value=0, max_value=6,
                                          tag="batch_mergecols_input", width=40, step=0)

            dpg.add_spacer(height=5)
            dpg.add_separator()

            # Results area (white background)
            with dpg.child_window(tag="batch_results_area") as results_panel:
                dpg.add_text("No results yet. Configure and run a batch experiment.", color=(100, 100, 100))

        # Apply white background theme to results area with dark text
        with dpg.theme() as results_theme:
            with dpg.theme_component(dpg.mvAll):
                dpg.add_theme_color(dpg.mvThemeCol_ChildBg, (255, 255, 255))
                dpg.add_theme_color(dpg.mvThemeCol_Text, (40, 40, 40))
                dpg.add_theme_color(dpg.mvThemeCol_FrameBg, (245, 245, 245))
                dpg.add_theme_color(dpg.mvThemeCol_Border, (200, 200, 200))
        dpg.bind_item_theme(results_panel, results_theme)

    # Initialize problem and algorithm lists
    _state["current_suite"] = initial_suites[0] if initial_suites else ""
    _update_prob_list_display()
    _update_algo_list_display()


def update():
    """Called each frame to poll progress."""
    status = _state.get("status")
    if status is None:
        return

    now = time.time()
    if status.running and (now - _state["last_poll"]) >= 1.0:
        _state["last_poll"] = now
        data_path = _state["file_manager"].get_data_path_str()
        completed = count_pkl_files(data_path)
        total = status.total_tasks
        pct = min(completed / total, 1.0) if total > 0 else 0.0
        elapsed = now - status.start_time if status.start_time else 0.0

        if dpg.does_item_exist("batch_progress"):
            dpg.set_value("batch_progress", pct)
        if dpg.does_item_exist("batch_elapsed_text"):
            dpg.set_value("batch_elapsed_text", f"Elapsed: {format_time(elapsed)}")
        if dpg.does_item_exist("batch_phase_text"):
            phase_text = {
                "running": f"Running experiments... ({completed}/{total})",
                "analyzing": "Analyzing results...",
                "complete": "Complete!",
            }.get(status.current_phase, status.current_phase)
            dpg.set_value("batch_phase_text", f"Phase: {phase_text}")
        if dpg.does_item_exist("batch_running_text"):
            dpg.set_value("batch_running_text",
                          f"Batch experiment: {completed}/{total} completed")

    elif not status.running and status.finished and not _state["displayed"]:
        _state["displayed"] = True
        # Unlock run button
        dpg.configure_item("batch_run_btn", enabled=True)
        if "run_theme" in _state:
            dpg.bind_item_theme("batch_run_btn", _state["run_theme"])
        dpg.delete_item("batch_results_area", children_only=True)

        if status.error:
            error_msg = f"Experiment failed: {status.error}"
            with dpg.group(horizontal=True, parent="batch_results_area"):
                dpg.add_text(error_msg, color=(200, 60, 60), wrap=760)
                dpg.add_button(label="Copy", callback=_on_copy_error_click,
                               user_data=error_msg, width=24)
        else:
            dpg.add_text(f"Experiment completed in {format_time(status.elapsed)}",
                         parent="batch_results_area", color=(60, 160, 60))
            save_path = _state.get("load_data_save_path") or _state["file_manager"].get_results_path_str()
            # Get merge_columns setting for display sizing (0 if merge unchecked)
            merge_checked = dpg.get_value("batch_merge_check") if dpg.does_item_exist("batch_merge_check") else False
            merge_cols = dpg.get_value("batch_mergecols_input") if (merge_checked and dpg.does_item_exist("batch_mergecols_input")) else 0
            _display_analysis_results(save_path, merge_cols)
