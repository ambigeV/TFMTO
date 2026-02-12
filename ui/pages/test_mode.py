"""Test Mode - Single-run optimization experiments."""

import os
import io
import time
import threading
import contextlib
import numpy as np
import dearpygui.dearpygui as dpg
from pathlib import Path

from utils.registry import (
    get_algorithm_names, get_algorithm_class,
    get_problem_suites, get_problem_methods, create_problem,
    is_multi_objective, get_problem_settings,
)
from utils.algo_scanner import (
    get_algorithm_params_from_scan, scan_all_algorithms,
    get_algorithm_info, format_algorithm_info,
)
from utils.problem_scanner import (
    get_scanned_problem_params, is_fixed_dimension_problem,
)
from utils.runner import RunStatus
from utils.file_manager import FileManager
from utils.backup_manager import BackupManager
from utils.settings_builder import SettingsBuilder
from config.default_params import PROBLEM_PARAMS, FIXED_DIMENSION_SUITES, FIXED_OBJECTIVES_SUITES
from config.constants import (
    CATEGORIES, METRICS, TABLE_FORMATS, FIGURE_FORMATS, STATISTIC_TYPES,
    COLOR_TITLE, COLOR_ERROR, COLOR_SUCCESS, COLOR_SECTION,
)
from components.dpg_helpers import (
    load_image_to_texture, show_error_modal, show_info_modal,
    show_confirm_dialog, format_time, get_texture_registry,
    copy_text_to_clipboard, _on_copy_error_click,
)

# Theme for reorder arrow buttons (created lazily)
_arrow_btn_theme = None

def _get_arrow_btn_theme():
    """Get or create the arrow button theme."""
    global _arrow_btn_theme
    if _arrow_btn_theme is None or not dpg.does_item_exist(_arrow_btn_theme):
        with dpg.theme() as _arrow_btn_theme:
            with dpg.theme_component(dpg.mvButton):
                dpg.add_theme_color(dpg.mvThemeCol_Button, (180, 180, 180))
                dpg.add_theme_color(dpg.mvThemeCol_ButtonHovered, (150, 150, 150))
                dpg.add_theme_color(dpg.mvThemeCol_ButtonActive, (120, 120, 120))
                dpg.add_theme_color(dpg.mvThemeCol_Text, (40, 40, 40))
    return _arrow_btn_theme


# Theme for disabled run button (created lazily)
_disabled_btn_theme = None

def _get_disabled_btn_theme():
    """Get or create the disabled button theme."""
    global _disabled_btn_theme
    if _disabled_btn_theme is None or not dpg.does_item_exist(_disabled_btn_theme):
        with dpg.theme() as _disabled_btn_theme:
            with dpg.theme_component(dpg.mvButton):
                dpg.add_theme_color(dpg.mvThemeCol_Button, (100, 100, 100))
                dpg.add_theme_color(dpg.mvThemeCol_ButtonHovered, (100, 100, 100))
                dpg.add_theme_color(dpg.mvThemeCol_ButtonActive, (100, 100, 100))
                dpg.add_theme_color(dpg.mvThemeCol_Text, (160, 160, 160))
    return _disabled_btn_theme


# Module state
_state = {
    "results": {},
    "statuses": {},
    "running": False,
    "file_manager": None,
    "backup_manager": None,
    "settings_builder": None,
    "current_prob_category": "STSO",
    "current_algo_category": "STSO",
    "current_suite": "",
    # Algorithm selection using unique IDs (not indices)
    "selected_algos": [],  # List of {id, algo_name, custom_name, params}
    "algo_id_counter": 0,  # Auto-increment ID for algorithms
    "algo_params_cache": {},  # {algo_id: {param_name: value}}
    "stdout_buffer": "",  # Captured stdout from worker thread
    # Problem info for analysis
    "run_prob_category": None,
    "run_suite": None,
    "run_method": None,
    "run_algos": [],
    "analysis_results": None,
}


def _init_managers(base_path: str):
    """Initialize file and backup managers."""
    _state["file_manager"] = FileManager(base_path)
    _state["backup_manager"] = BackupManager(base_path)
    _state["settings_builder"] = SettingsBuilder()
    _state["file_manager"].ensure_structure()

    # Pre-scan algorithms on startup
    scan_all_algorithms()


def _on_prob_category_change(sender, app_data):
    """Update suite combo when problem category changes."""
    cat = app_data
    _state["current_prob_category"] = cat
    suites = get_problem_suites(cat)
    dpg.configure_item("test_suite_combo", items=suites)
    if suites:
        dpg.set_value("test_suite_combo", suites[0])
        _on_suite_change(None, suites[0])

    _update_problem_params_visibility(cat, suites[0] if suites else "")
    _update_metric_visibility(cat)


def _on_algo_category_change(sender, app_data):
    """Update algorithm list when algorithm category changes."""
    cat = app_data
    _state["current_algo_category"] = cat
    _state["selected_algos"] = []
    _state["algo_id_counter"] = 0
    _state["algo_params_cache"] = {}
    _update_algo_list_display()
    _update_algo_params_display()


def _on_suite_change(sender, app_data):
    """Update method combo when suite changes."""
    cat = dpg.get_value("test_prob_cat_combo")
    suite = app_data
    _state["current_suite"] = suite
    methods = get_problem_methods(cat, suite)
    dpg.configure_item("test_method_combo", items=methods)
    if methods:
        dpg.set_value("test_method_combo", methods[0])

    _update_problem_params_visibility(cat, suite)


def _update_problem_params_visibility(cat: str, suite: str):
    """Show/hide problem parameters based on suite (auto-scanned or hardcoded)."""
    # Try auto-scanned params first, fall back to hardcoded
    scanned_params = get_scanned_problem_params(cat, suite)
    suite_params = scanned_params if scanned_params else PROBLEM_PARAMS.get(suite, {})

    # Use scanner to determine if fixed dimension, fall back to hardcoded list
    is_fixed_dim = is_fixed_dimension_problem(cat, suite) if scanned_params else (suite in FIXED_DIMENSION_SUITES)

    # Check if objectives are fixed (e.g., ZDT, CF, UF have fixed M per problem)
    is_fixed_obj = suite in FIXED_OBJECTIVES_SUITES

    # Show D only if suite is not fixed dimension
    show_D = "D" in suite_params or not is_fixed_dim
    # Show M only if it's in params AND not a fixed-objective suite
    show_M = "M" in suite_params and not is_fixed_obj
    show_K = "K" in suite_params
    show_L = "L" in suite_params

    # Show params group if any parameter should be visible
    has_any_params = show_D or show_M or show_K or show_L
    if dpg.does_item_exist("test_problem_params_group"):
        dpg.configure_item("test_problem_params_group", show=has_any_params)

    if dpg.does_item_exist("test_D_group"):
        dpg.configure_item("test_D_group", show=show_D)

    if dpg.does_item_exist("test_M_group"):
        dpg.configure_item("test_M_group", show=show_M)

    if dpg.does_item_exist("test_K_group"):
        dpg.configure_item("test_K_group", show=show_K)

    if dpg.does_item_exist("test_L_group"):
        dpg.configure_item("test_L_group", show=show_L)


def _update_metric_visibility(cat: str):
    """No-op: metric combo is always visible now."""
    pass


def _on_algo_click(sender, app_data, user_data):
    """Handle algorithm button click to add to selected list."""
    algo_name = user_data
    # Create new algorithm entry with unique ID
    algo_id = _state["algo_id_counter"]
    _state["algo_id_counter"] += 1
    _state["selected_algos"].append({
        "id": algo_id,
        "algo_name": algo_name,
        "custom_name": algo_name,  # Default to algorithm name
    })
    _update_algo_params_display()


def _remove_algo(sender, app_data, user_data):
    """Remove algorithm from selected list by ID."""
    algo_id = user_data
    _state["selected_algos"] = [a for a in _state["selected_algos"] if a["id"] != algo_id]
    # Clean up cached params for this algorithm
    if algo_id in _state["algo_params_cache"]:
        del _state["algo_params_cache"][algo_id]
    _update_algo_params_display()


def _update_algo_list_display():
    """Update the algorithm list with clickable buttons and tooltips."""
    if not dpg.does_item_exist("test_algo_list_container"):
        return

    dpg.delete_item("test_algo_list_container", children_only=True)

    cat = _state["current_algo_category"]
    algos = get_algorithm_names(cat)

    # Labels for algorithm info
    info_labels = {
        'n_tasks': 'Tasks', 'dims': 'Dimensions', 'objs': 'Objectives',
        'n_objs': 'Num Objectives', 'cons': 'Constraints', 'n_cons': 'Num Constraints',
        'expensive': 'Expensive', 'knowledge_transfer': 'Knowledge Transfer',
        'n': 'Population', 'max_nfes': 'Max NFEs',
    }

    for algo_name in algos:
        btn = dpg.add_button(label=algo_name, callback=_on_algo_click, user_data=algo_name,
                             width=-1, parent="test_algo_list_container")
        # Add tooltip with algorithm information (blue labels)
        info = get_algorithm_info(cat, algo_name)
        if info:
            with dpg.tooltip(btn):
                for key, value in info.items():
                    label = info_labels.get(key, key)
                    with dpg.group(horizontal=True):
                        dpg.add_text(f"{label}:", color=COLOR_TITLE)
                        dpg.add_text(f" {value}")


def _move_algo_up(sender, app_data, user_data):
    """Move algorithm up in the selected list by ID."""
    algo_id = user_data
    selected = _state["selected_algos"]
    # Find index of algorithm with this ID
    idx = next((i for i, a in enumerate(selected) if a["id"] == algo_id), -1)
    if idx <= 0 or idx >= len(selected):
        return
    selected[idx], selected[idx - 1] = selected[idx - 1], selected[idx]
    _update_algo_params_display()


def _move_algo_down(sender, app_data, user_data):
    """Move algorithm down in the selected list by ID."""
    algo_id = user_data
    selected = _state["selected_algos"]
    # Find index of algorithm with this ID
    idx = next((i for i, a in enumerate(selected) if a["id"] == algo_id), -1)
    if idx < 0 or idx >= len(selected) - 1:
        return
    selected[idx], selected[idx + 1] = selected[idx + 1], selected[idx]
    _update_algo_params_display()


def _get_algo_source_path(algo_name: str) -> Path:
    """Get the source file path for an algorithm."""
    cat = _state["current_algo_category"]

    # Get file name from display name
    display_to_file = {
        'CMA-ES': 'CMA_ES', 'MA-ES': 'MA_ES', 'IPOP-CMA-ES': 'IPOP_CMA_ES',
        'sep-CMA-ES': 'sep_CMA_ES', 'OpenAI-ES': 'OpenAI_ES', 'SA-COSO': 'SA_COSO',
        'GL-SADE': 'GL_SADE', 'KL-PSO': 'KL_PSO', 'SL-PSO': 'SL_PSO', 'EEI-BO': 'EEI_BO',
        'NSGA-II': 'NSGA_II', 'NSGA-III': 'NSGA_III', 'NSGA-II-SDR': 'NSGA_II_SDR',
        'MOEA/D': 'MOEA_D', 'MOEA/D-STM': 'MOEA_D_STM', 'MOEA/D-FRRMAB': 'MOEA_D_FRRMAB',
        'MOEA/DD': 'MOEA_DD', 'C-TAEA': 'C_TAEA', 'K-RVEA': 'K_RVEA',
        'DSAEA-PS': 'DSAEA_PS', 'MCEA-D': 'MCEA_D', 'CPS-MOEA': 'CPS_MOEA',
        'MFEA-II': 'MFEA_II', 'G-MFEA': 'G_MFEA', 'MTEA-AD': 'MTEA_AD',
        'MTEA-SaO': 'MTEA_SaO', 'LCB-EMT': 'LCB_EMT', 'EEI-BO+': 'EEI_BO_plus',
        'MO-MFEA': 'MO_MFEA', 'MO-MFEA-II': 'MO_MFEA_II', 'MO-EMEA': 'MO_EMEA',
        'EMT-ET': 'EMT_ET', 'EMT-PD': 'EMT_PD', 'MTDE-MKTA': 'MTDE_MKTA',
        'MO-MTEA-SaO': 'MO_MTEA_SaO', 'ParEGO-KT': 'ParEGO_KT', 'MTEA-D-DN': 'MTEA_D_DN',
    }
    file_name = display_to_file.get(algo_name, algo_name)

    ui_dir = Path(__file__).resolve().parent.parent
    project_root = ui_dir.parent
    return project_root / 'src' / 'ddmtolab' / 'Algorithms' / cat / f'{file_name}.py'


def _open_algo_source(sender, app_data, user_data):
    """Open algorithm source file in PyCharm or default editor."""
    import subprocess
    import sys
    import shutil

    algo_name = user_data
    source_path = _get_algo_source_path(algo_name)

    if not source_path.exists():
        show_error_modal(f"Source file not found:\n{source_path}")
        return

    # Try PyCharm first
    pycharm_paths = [
        shutil.which('pycharm'),
        shutil.which('pycharm64'),
        r'C:\Program Files\JetBrains\PyCharm Community Edition 2024.1\bin\pycharm64.exe',
        r'C:\Program Files\JetBrains\PyCharm 2024.1\bin\pycharm64.exe',
        r'C:\Program Files\JetBrains\PyCharm Community Edition 2023.3\bin\pycharm64.exe',
        r'C:\Program Files\JetBrains\PyCharm 2023.3\bin\pycharm64.exe',
    ]

    # Find PyCharm installation
    pycharm_exe = None
    for path in pycharm_paths:
        if path and Path(path).exists():
            pycharm_exe = path
            break

    # Also check common JetBrains toolbox location
    if not pycharm_exe:
        toolbox_base = Path.home() / 'AppData' / 'Local' / 'JetBrains' / 'Toolbox' / 'apps'
        if toolbox_base.exists():
            for pycharm_dir in toolbox_base.glob('PyCharm*/**/pycharm64.exe'):
                pycharm_exe = str(pycharm_dir)
                break

    if pycharm_exe:
        # Open in PyCharm
        subprocess.Popen([pycharm_exe, str(source_path)],
                         creationflags=subprocess.CREATE_NO_WINDOW if sys.platform == 'win32' else 0)
    else:
        # Fallback to default editor
        if sys.platform == 'win32':
            os.startfile(str(source_path))
        elif sys.platform == 'darwin':
            subprocess.run(['open', str(source_path)])
        else:
            subprocess.run(['xdg-open', str(source_path)])


def _update_algo_params_display():
    """Update algorithm parameters display based on selection."""
    if not dpg.does_item_exist("test_algo_params_container"):
        return

    dpg.delete_item("test_algo_params_container", children_only=True)

    selected = _state["selected_algos"]
    cat = _state["current_algo_category"]

    if not selected:
        return

    # Each algorithm as a collapsible header with controls
    for i, algo_entry in enumerate(selected):
        algo_id = algo_entry["id"]
        algo_name = algo_entry["algo_name"]
        custom_name = algo_entry["custom_name"]

        params = get_algorithm_params_from_scan(cat, algo_name)
        safe_algo_name = algo_name.replace("/", "-")

        # Header shows custom name
        header_label = custom_name

        # Up/Down buttons for reordering
        with dpg.group(horizontal=True, parent="test_algo_params_container"):
            dpg.add_button(label="^", width=20, callback=_move_algo_up, user_data=algo_id,
                          enabled=(i > 0))
            dpg.add_button(label="v", width=20, callback=_move_algo_down, user_data=algo_id,
                          enabled=(i < len(selected) - 1))

        # Collapsing header with algorithm name
        with dpg.collapsing_header(label=f"{header_label}##{safe_algo_name}_{algo_id}",
                                   parent="test_algo_params_container",
                                   default_open=(len(selected) == 1)) as header:
            # Right-click popup menu on header
            with dpg.popup(header, mousebutton=dpg.mvMouseButton_Right,
                           min_size=(160, 0), max_size=(160, 50)):
                dpg.add_menu_item(label="Open Source File",
                                  callback=_open_algo_source, user_data=algo_name)

            # Name parameter for renaming
            with dpg.group(horizontal=True):
                dpg.add_text("Name:")
                name_tag = f"test_algo_name_{algo_id}"
                dpg.add_input_text(tag=name_tag, default_value=custom_name, width=-1,
                                   callback=lambda s, a, u: _on_algo_name_change(u, a),
                                   user_data=algo_id)

            # Other parameters
            for param_name, param_info in params.items():
                _add_param_input_with_id(algo_id, algo_name, param_name, param_info)

            # Delete button at the end
            dpg.add_spacer(height=5)
            del_btn = dpg.add_button(label="Delete Algorithm", callback=_remove_algo, user_data=algo_id,
                          width=-1)
            if _state.get("delete_theme"):
                dpg.bind_item_theme(del_btn, _state["delete_theme"])


def _on_algo_name_change(algo_id: int, new_name: str):
    """Handle algorithm name change by ID."""
    for algo_entry in _state["selected_algos"]:
        if algo_entry["id"] == algo_id:
            algo_entry["custom_name"] = new_name
            break


def _add_param_input_with_id(algo_id: int, algo_name: str, param_name: str, param_info: dict):
    """Add a parameter input widget using algorithm ID for tracking."""
    param_type = param_info.get('type', 'int')
    default = param_info.get('default')
    desc = param_name

    # Get cached value if exists
    if algo_id in _state.get("algo_params_cache", {}):
        cached = _state["algo_params_cache"][algo_id]
        if param_name in cached:
            default = cached[param_name]

    tag = f"test_param_{algo_id}_{param_name}"

    # Check if parameter is vectorizable based on algorithm_information
    cat = _state["current_algo_category"]
    algo_info = get_algorithm_info(cat, algo_name)
    is_vectorizable = algo_info.get(param_name, '') == 'unequal'

    # Use closure to capture values properly
    def make_callback(aid, pn):
        return lambda s, a: _cache_param_value_by_id(aid, pn, a)

    with dpg.group(horizontal=True):
        if is_vectorizable:
            dpg.add_text(f"{desc}:", tag=f"{tag}_label", color=COLOR_SECTION)
            # Use text input for vectorizable parameters to allow list input
            default_str = str(default) if default is not None else "100" if param_type == 'int' else "0.5"
            dpg.add_input_text(tag=tag, default_value=default_str, width=-1, hint="e.g. [100,200]",
                               callback=make_callback(algo_id, param_name))
        else:
            dpg.add_text(f"{desc}:", tag=f"{tag}_label")
            if param_type == 'int':
                dpg.add_input_int(tag=tag, default_value=default if default is not None else 100,
                                  width=-1, step=0, callback=make_callback(algo_id, param_name))
            elif param_type == 'float':
                dpg.add_input_float(tag=tag, default_value=default if default is not None else 0.5,
                                    width=-1, step=0, format="%.4f",
                                    callback=make_callback(algo_id, param_name))
            elif param_type == 'str':
                dpg.add_input_text(tag=tag, default_value=default if default else "",
                                   width=-1, callback=make_callback(algo_id, param_name))
            elif param_type == 'bool':
                dpg.add_checkbox(tag=tag, default_value=default if default is not None else False,
                                 callback=make_callback(algo_id, param_name))
            else:
                dpg.add_input_text(tag=tag, default_value=str(default) if default else "",
                                   width=-1, callback=make_callback(algo_id, param_name))


def _cache_param_value_by_id(algo_id: int, param_name: str, value):
    """Cache parameter value for persistence using algorithm ID."""
    if "algo_params_cache" not in _state:
        _state["algo_params_cache"] = {}
    if algo_id not in _state["algo_params_cache"]:
        _state["algo_params_cache"][algo_id] = {}
    _state["algo_params_cache"][algo_id][param_name] = value


def _parse_param_value(value, param_type: str):
    """Parse parameter value, supporting list input like [100, 200]."""
    if isinstance(value, str):
        value = value.strip()
        # Check if it's a list format
        if value.startswith('[') and value.endswith(']'):
            try:
                import ast
                parsed = ast.literal_eval(value)
                if isinstance(parsed, list):
                    # Convert list elements to appropriate type
                    if param_type == 'int':
                        return [int(x) for x in parsed]
                    elif param_type == 'float':
                        return [float(x) for x in parsed]
                    return parsed
            except:
                pass
        # Single value - convert to appropriate type
        try:
            if param_type == 'int':
                return int(value)
            elif param_type == 'float':
                return float(value)
        except:
            pass
    return value


def _get_algo_params_for_run(algo_id: int, algo_name: str) -> dict:
    """Get parameters for a specific algorithm by ID."""
    params = {}
    cat = _state["current_algo_category"]
    algo_params = get_algorithm_params_from_scan(cat, algo_name)

    for param_name, param_info in algo_params.items():
        tag = f"test_param_{algo_id}_{param_name}"
        if dpg.does_item_exist(tag):
            value = dpg.get_value(tag)
            # Parse value (handles list input for vectorizable params)
            params[param_name] = _parse_param_value(value, param_info.get('type', 'int'))

    return params


def _select_all_algos(sender, app_data):
    """Select all algorithms from current category."""
    cat = _state["current_algo_category"]
    algos = get_algorithm_names(cat)
    _state["selected_algos"] = []
    for algo_name in algos:
        algo_id = _state["algo_id_counter"]
        _state["algo_id_counter"] += 1
        _state["selected_algos"].append({
            "id": algo_id,
            "algo_name": algo_name,
            "custom_name": algo_name,
        })
    _update_algo_params_display()


def _deselect_all_algos(sender, app_data):
    """Clear all selected algorithms."""
    _state["selected_algos"] = []
    _state["algo_params_cache"] = {}
    _update_algo_params_display()


def _run_clicked(sender, app_data):
    """Run selected algorithms sequentially."""
    if _state["running"]:
        return

    prob_cat = dpg.get_value("test_prob_cat_combo")
    algo_cat = dpg.get_value("test_algo_cat_combo")
    suite = dpg.get_value("test_suite_combo")
    method = dpg.get_value("test_method_combo")

    # Get problem parameters (use scanner to determine if fixed dimension)
    problem_kwargs = {}
    is_fixed_dim = is_fixed_dimension_problem(prob_cat, suite)
    # D is only available if suite is not fixed dimension
    if not is_fixed_dim:
        if dpg.does_item_exist("test_D_input") and dpg.is_item_shown("test_D_group"):
            problem_kwargs["D"] = dpg.get_value("test_D_input")
    # M, K, L are available based on suite params (independent of fixed dimension)
    if dpg.does_item_exist("test_M_input") and dpg.is_item_shown("test_M_group"):
        problem_kwargs["M"] = dpg.get_value("test_M_input")
    if dpg.does_item_exist("test_K_input") and dpg.is_item_shown("test_K_group"):
        problem_kwargs["K"] = dpg.get_value("test_K_input")
    if dpg.does_item_exist("test_L_input") and dpg.is_item_shown("test_L_group"):
        problem_kwargs["L"] = dpg.get_value("test_L_input")

    # Get selected algorithms in user-defined order
    selected = _state["selected_algos"]  # List of {id, algo_name, custom_name}

    if not selected:
        show_error_modal("Please select at least one algorithm.")
        return

    # Check for duplicate custom names
    custom_names = [a["custom_name"] for a in selected]
    if len(custom_names) != len(set(custom_names)):
        show_error_modal("Duplicate algorithm names detected.\nPlease rename algorithms to have unique names.")
        return

    # Store problem info for analysis - use custom names (replace "/" with "-")
    _state["run_prob_category"] = prob_cat
    _state["run_suite"] = suite
    _state["run_method"] = method
    _state["run_algos"] = [a["custom_name"].replace("/", "-") for a in selected]  # Safe names in user order
    _state["analysis_results"] = None

    _state["results"] = {}
    _state["statuses"] = {}
    _state["running"] = True
    _state["results_displayed"] = False

    # Lock run button
    dpg.configure_item("test_run_btn", enabled=False)
    dpg.bind_item_theme("test_run_btn", _get_disabled_btn_theme())

    # Clear previous results
    dpg.delete_item("test_results_area", children_only=True)

    # Show progress bar and elapsed time
    _state["stdout_buffer"] = ""
    dpg.add_text("Running...", parent="test_results_area", tag="test_running_text", color=(60, 60, 60))
    dpg.add_progress_bar(tag="test_progress", parent="test_results_area",
                         default_value=0.0, width=-1)
    dpg.add_text("Elapsed: 0s", parent="test_results_area", tag="test_elapsed_text", color=(100, 100, 100))

    # Create problem
    try:
        problem = create_problem(prob_cat, suite, method, **problem_kwargs)
    except Exception as e:
        show_error_modal(f"Failed to create problem: {e}")
        _state["running"] = False
        dpg.configure_item("test_run_btn", enabled=True)
        dpg.configure_item("test_stop_btn", enabled=False)
        return

    save_path = _state["file_manager"].get_data_path_str()

    def _worker():
        capture = io.StringIO()
        start_time = time.time()
        with contextlib.redirect_stdout(capture), contextlib.redirect_stderr(capture):
            for i, algo_entry in enumerate(selected):
                if not _state["running"]:
                    break

                algo_id = algo_entry["id"]
                algo_name = algo_entry["algo_name"]
                custom_name = algo_entry["custom_name"]
                display_name = custom_name  # Use custom name for display

                if dpg.does_item_exist("test_running_text"):
                    dpg.set_value("test_running_text", f"Running {display_name} ({i+1}/{len(selected)})...")
                if dpg.does_item_exist("test_progress"):
                    dpg.set_value("test_progress", i / len(selected))

                try:
                    algo_cls = get_algorithm_class(algo_cat, algo_name)

                    # Get algorithm-specific parameters using ID
                    algo_params = _get_algo_params_for_run(algo_id, algo_name)

                    # Build algorithm kwargs
                    # Replace "/" with "-" in custom name for file saving
                    safe_name = custom_name.replace("/", "-")
                    algo_kwargs = {
                        "problem": problem,
                        "save_data": True,
                        "save_path": save_path,
                        "name": safe_name,  # Use safe custom name for saving
                        "disable_tqdm": True,
                    }

                    # Add algorithm-specific params (filter out None values)
                    for k, v in algo_params.items():
                        if v is not None:
                            algo_kwargs[k] = v

                    algo_instance = algo_cls(**algo_kwargs)

                    status = RunStatus()
                    status.running = True
                    status.start_time = time.time()
                    try:
                        result = algo_instance.optimize()
                        # Handle algorithms that return tuple (Results, extra_data)
                        if isinstance(result, tuple):
                            result = result[0]
                        status.result = result
                        status.finished = True
                    except Exception as e:
                        status.error = str(e)
                        status.finished = True
                    finally:
                        status.running = False
                        status.elapsed = time.time() - status.start_time

                    _state["statuses"][display_name] = status
                    if status.result is not None:
                        _state["results"][display_name] = status.result

                except Exception as e:
                    import traceback
                    _state["statuses"][display_name] = RunStatus(error=f"{e}\n{traceback.format_exc()}", finished=True)

                # Flush captured output to buffer
                _state["stdout_buffer"] = capture.getvalue()

                # Update elapsed time
                if dpg.does_item_exist("test_elapsed_text"):
                    elapsed = time.time() - start_time
                    dpg.set_value("test_elapsed_text", f"Elapsed: {format_time(elapsed)}")

        # Final flush
        _state["stdout_buffer"] = capture.getvalue()
        if dpg.does_item_exist("test_progress"):
            dpg.set_value("test_progress", 1.0)
        _state["running"] = False

    t = threading.Thread(target=_worker, daemon=True)
    t.start()


def _stop_clicked(sender, app_data):
    """Stop running algorithms."""
    if _state["running"]:
        _state["running"] = False
        if dpg.does_item_exist("test_running_text"):
            dpg.set_value("test_running_text", "Stopped by user")
            dpg.configure_item("test_running_text", color=(150, 100, 50))


def _clean_clicked(sender, app_data):
    """Show cleanup confirmation dialog."""
    def on_backup_and_clean():
        backup_path = _state["backup_manager"].clean_and_backup()
        if backup_path:
            show_info_modal(f"Backup created at:\n{backup_path}\n\nData cleaned.")
        else:
            show_info_modal("No data to backup. Folders cleaned.")
        dpg.delete_item("test_results_area", children_only=True)
        dpg.add_text("Data cleaned. Ready for new experiment.", parent="test_results_area",
                     color=(100, 100, 100))

    def on_clean_only():
        _state["backup_manager"].clean_without_backup()
        show_info_modal("Data cleaned without backup.")
        dpg.delete_item("test_results_area", children_only=True)
        dpg.add_text("Data cleaned. Ready for new experiment.", parent="test_results_area",
                     color=(100, 100, 100))

    show_confirm_dialog(
        "Do you want to backup data before cleaning?",
        on_yes=on_backup_and_clean,
        on_no=on_clean_only,
        title="Clean Data",
        yes_label="Backup & Clean",
        no_label="Clean Only",
        cancel_label="Cancel"
    )


def _run_analysis():
    """Run TestDataAnalyzer to generate PNG figures."""
    prob_cat = _state["run_prob_category"]
    suite = _state["run_suite"]
    method = _state["run_method"]
    algo_order = _state["run_algos"]

    try:
        from ddmtolab.Methods.test_data_analysis import TestDataAnalyzer

        data_path = _state["file_manager"].get_data_path_str()
        results_path = _state["file_manager"].get_results_path_str()

        data_files = list(Path(data_path).glob("*.pkl"))
        if not data_files:
            return None

        settings = None
        if is_multi_objective(prob_cat):
            settings = get_problem_settings(prob_cat, suite)
            if settings:
                settings = settings.copy()
                settings['problems'] = [method]
                metric = dpg.get_value("test_metric_combo") if dpg.does_item_exist("test_metric_combo") else "IGD"
                settings['metric'] = metric

        use_log = dpg.get_value("test_log_check") if dpg.does_item_exist("test_log_check") else False
        show_pf = dpg.get_value("test_showpf_check") if dpg.does_item_exist("test_showpf_check") else True
        show_nd = dpg.get_value("test_shownd_check") if dpg.does_item_exist("test_shownd_check") else True
        best_so_far = dpg.get_value("test_bestsofar_check") if dpg.does_item_exist("test_bestsofar_check") else True
        figure_format = dpg.get_value("test_fig_combo") if dpg.does_item_exist("test_fig_combo") else "png"

        analyzer = TestDataAnalyzer(
            data_path=data_path,
            settings=settings,
            save_path=results_path,
            algorithm_order=algo_order,
            figure_format=figure_format,
            log_scale=use_log,
            show_pf=show_pf,
            show_nd=show_nd,
            best_so_far=best_so_far,
            clear_results=True
        )

        return analyzer.run()

    except Exception:
        return None


def _open_file_location(sender, app_data, user_data):
    """Open the folder containing the file in Windows Explorer."""
    import subprocess
    file_path = user_data
    folder_path = os.path.dirname(file_path)
    if os.path.exists(folder_path):
        subprocess.Popen(['explorer', '/select,', file_path.replace('/', '\\')])


def _copy_image_to_clipboard(sender, app_data, user_data):
    """Copy image to clipboard (Windows only)."""
    import sys
    file_path = user_data

    if not os.path.exists(file_path):
        return

    try:
        from PIL import Image
        import io

        if sys.platform == 'win32':
            import win32clipboard

            # Open image and convert to BMP format for clipboard
            img = Image.open(file_path)
            if img.mode != 'RGB':
                img = img.convert('RGB')

            # Save to bytes as BMP
            output = io.BytesIO()
            img.save(output, format='BMP')
            bmp_data = output.getvalue()[14:]  # Remove BMP header
            output.close()

            # Copy to clipboard
            win32clipboard.OpenClipboard()
            win32clipboard.EmptyClipboard()
            win32clipboard.SetClipboardData(win32clipboard.CF_DIB, bmp_data)
            win32clipboard.CloseClipboard()
        else:
            # For non-Windows, just show a message
            show_info_modal("Copy to clipboard is only supported on Windows.")
    except ImportError:
        show_error_modal("Please install pywin32: pip install pywin32")
    except Exception as e:
        show_error_modal(f"Failed to copy image: {e}")


def _load_and_display_image(image_path: str, parent, max_width: int = 600):
    """Load a PNG image and display it in the UI."""
    try:
        from PIL import Image
        import numpy as np

        if not os.path.exists(image_path):
            return False

        img = Image.open(image_path).convert("RGBA")
        w, h = img.size

        if w > max_width:
            scale = max_width / w
            w = max_width
            h = int(h * scale)
            img = img.resize((w, h), Image.Resampling.LANCZOS)

        data = np.array(img).astype(np.float32) / 255.0
        flat = data.flatten().tolist()

        tex_reg = get_texture_registry()
        if not dpg.does_item_exist(tex_reg):
            return False

        tex_tag = dpg.add_static_texture(
            width=w, height=h, default_value=flat,
            parent=tex_reg
        )

        if parent:
            if not dpg.does_item_exist(parent):
                return False
            img_widget = dpg.add_image(tex_tag, width=w, height=h, parent=parent)
        else:
            img_widget = dpg.add_image(tex_tag, width=w, height=h)

        with dpg.popup(img_widget, mousebutton=dpg.mvMouseButton_Right):
            dpg.add_menu_item(label="Copy Image",
                              callback=_copy_image_to_clipboard, user_data=image_path)
            dpg.add_menu_item(label="Open File Location",
                              callback=_open_file_location, user_data=image_path)

        return True

    except Exception:
        return False


def _display_results():
    """Build results display showing Excel tables and PNG images from TestDataAnalyzer."""
    results = _state["results"]
    statuses = _state["statuses"]

    dpg.delete_item("test_results_area", children_only=True)

    if not results and not statuses:
        dpg.add_text("No results yet. Configure and run an experiment.", parent="test_results_area",
                     color=(100, 100, 100))
        return

    for name, status in statuses.items():
        if status.error:
            error_msg = f"[ERROR] {name}: {status.error}"
            with dpg.group(horizontal=True, parent="test_results_area"):
                dpg.add_text(error_msg, color=(200, 60, 60), wrap=560)
                dpg.add_button(label="\u2398", callback=_on_copy_error_click,
                               user_data=error_msg, width=24)

    if not results:
        return

    dpg.add_text("Generating analysis...", parent="test_results_area", tag="test_analysis_status",
                 color=(60, 60, 60))
    _run_analysis()

    if dpg.does_item_exist("test_analysis_status"):
        dpg.delete_item("test_analysis_status")

    results_path = Path(_state["file_manager"].get_results_path_str())

    # Image layout: 3 images per row, uniform width
    viewport_w = dpg.get_viewport_width()
    available_width = max(400, viewport_w - 250 - 280 - 50)
    IMG_WIDTH = (available_width - 40) // 3  # 3 images + spacing per row
    nd_per_row = 3

    # Open Results Folder button (like batch mode)
    def _open_folder():
        import subprocess
        import sys
        if sys.platform == 'win32':
            os.startfile(str(results_path.resolve()))
        elif sys.platform == 'darwin':
            subprocess.run(['open', str(results_path.resolve())])
        else:
            subprocess.run(['xdg-open', str(results_path.resolve())])

    with dpg.group(horizontal=True, parent="test_results_area"):
        dpg.add_spacer(width=-1)
        btn = dpg.add_button(label="Open Results Folder", callback=lambda: _open_folder(), width=150)
        with dpg.theme() as btn_theme:
            with dpg.theme_component(dpg.mvButton):
                dpg.add_theme_color(dpg.mvThemeCol_Text, (40, 40, 40))
                dpg.add_theme_color(dpg.mvThemeCol_Button, (200, 200, 200))
                dpg.add_theme_color(dpg.mvThemeCol_ButtonHovered, (180, 180, 180))
                dpg.add_theme_color(dpg.mvThemeCol_ButtonActive, (160, 160, 160))
        dpg.bind_item_theme(btn, btn_theme)

    dpg.add_spacer(height=5, parent="test_results_area")

    # Display Excel table first (like batch mode)
    excel_files = list(results_path.glob("*.xlsx"))
    if excel_files:
        dpg.add_text("Results Table", parent="test_results_area", color=COLOR_TITLE)
        dpg.add_separator(parent="test_results_area")

        # Create table theme
        with dpg.theme() as table_theme:
            with dpg.theme_component(dpg.mvTable):
                dpg.add_theme_color(dpg.mvThemeCol_TableHeaderBg, (70, 70, 70))
                dpg.add_theme_color(dpg.mvThemeCol_Text, (255, 255, 255))
            with dpg.theme_component(dpg.mvTableRow):
                dpg.add_theme_color(dpg.mvThemeCol_Text, (40, 40, 40))

        for excel_file in excel_files:
            try:
                from openpyxl import load_workbook
                import pandas as pd

                wb = load_workbook(excel_file)
                ws = wb.active

                # Build bold cells set
                bold_cells = set()
                for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=1):
                    for col_idx, cell in enumerate(row, start=0):
                        if cell.font and cell.font.bold:
                            bold_cells.add((row_idx, col_idx))

                df = pd.read_excel(excel_file)
                df = df.loc[:, ~df.columns.str.contains('^Unnamed')]
                df = df.dropna(axis=1, how='all')
                cols = list(df.columns)

                tbl = dpg.add_table(parent="test_results_area", header_row=True,
                                    borders_innerH=True, borders_outerH=True,
                                    borders_innerV=True, borders_outerV=True,
                                    policy=dpg.mvTable_SizingStretchSame,
                                    no_host_extendX=True, precise_widths=True)
                dpg.bind_item_theme(tbl, table_theme)

                for i, col in enumerate(cols):
                    if i == 0:
                        col_width = 150
                    elif i == 1:
                        col_width = 50
                    else:
                        col_width = 180
                    dpg.add_table_column(label=str(col), parent=tbl, width_fixed=True, init_width_or_weight=col_width)

                for row_idx, row in df.iterrows():
                    with dpg.table_row(parent=tbl):
                        for col_idx, col in enumerate(cols):
                            cell_text = str(row[col])
                            is_bold = (row_idx + 1, col_idx) in bold_cells
                            if is_bold:
                                dpg.add_text(cell_text, color=COLOR_TITLE)
                            else:
                                dpg.add_text(cell_text)
            except Exception as e:
                dpg.add_text(f"Could not load table: {e}", parent="test_results_area", color=(200, 60, 60))

        dpg.add_spacer(height=10, parent="test_results_area")

    # Display PNG images
    png_files = list(results_path.glob("*.png"))

    if png_files:
        convergence_files = [f for f in png_files if "convergence" in f.name.lower()]
        runtime_files = [f for f in png_files if "runtime" in f.name.lower()]
        other_files = [f for f in png_files if f not in convergence_files and f not in runtime_files]

        # 1. Convergence curves (adaptive per row, same width as all images)
        if convergence_files:
            dpg.add_text("Convergence Curves", parent="test_results_area", color=COLOR_TITLE)
            dpg.add_separator(parent="test_results_area")
            for i in range(0, len(convergence_files), nd_per_row):
                row_files = sorted(convergence_files)[i:i+nd_per_row]
                with dpg.group(horizontal=True, parent="test_results_area"):
                    for png_file in row_files:
                        dpg.add_spacer(height=5)
                        _load_and_display_image(str(png_file), 0, max_width=IMG_WIDTH)
                        dpg.add_spacer(width=10)
                dpg.add_spacer(height=8, parent="test_results_area")

        # 2. Non-dominated solutions (adaptive per row)
        nd_folder = results_path / "ND_Solutions"
        if nd_folder.exists():
            nd_pngs = sorted(nd_folder.glob("*.png"))
            if nd_pngs:
                dpg.add_spacer(height=10, parent="test_results_area")
                dpg.add_text("Non-Dominated Solutions", parent="test_results_area", color=COLOR_TITLE)
                dpg.add_separator(parent="test_results_area")
                dpg.add_spacer(height=5, parent="test_results_area")

                for i in range(0, len(nd_pngs), nd_per_row):
                    row_files = nd_pngs[i:i+nd_per_row]
                    with dpg.group(horizontal=True, parent="test_results_area"):
                        for png_file in row_files:
                            _load_and_display_image(str(png_file), 0, max_width=IMG_WIDTH)
                            dpg.add_spacer(width=10)
                    dpg.add_spacer(height=8, parent="test_results_area")

        # 3. Runtime comparison (adaptive per row)
        if runtime_files:
            dpg.add_spacer(height=10, parent="test_results_area")
            dpg.add_text("Runtime Comparison", parent="test_results_area", color=COLOR_TITLE)
            dpg.add_separator(parent="test_results_area")
            for i in range(0, len(runtime_files), nd_per_row):
                row_files = sorted(runtime_files)[i:i+nd_per_row]
                with dpg.group(horizontal=True, parent="test_results_area"):
                    for png_file in row_files:
                        dpg.add_spacer(height=5)
                        _load_and_display_image(str(png_file), 0, max_width=IMG_WIDTH)
                        dpg.add_spacer(width=10)
                dpg.add_spacer(height=8, parent="test_results_area")

        # 4. Other files
        for png_file in sorted(other_files):
            dpg.add_spacer(height=5, parent="test_results_area")
            _load_and_display_image(str(png_file), "test_results_area", max_width=IMG_WIDTH)
            dpg.add_spacer(height=10, parent="test_results_area")

    if not png_files and not excel_files:
        dpg.add_text("No results generated.", parent="test_results_area", color=(180, 80, 80))

    dpg.add_spacer(height=10, parent="test_results_area")
    dpg.add_text(f"Results saved to: {results_path}", parent="test_results_area", color=(100, 100, 100))


def _animation_clicked(sender, app_data):
    """Show animation settings dialog."""
    modal_tag = "animation_modal"
    if dpg.does_item_exist(modal_tag):
        dpg.delete_item(modal_tag)

    with dpg.window(label="Animation Generation", modal=True, tag=modal_tag, no_close=True,
                    width=360, autosize=True, pos=[500, 250]):
        dpg.add_spacer(height=5)
        with dpg.group(horizontal=True):
            dpg.add_text("Title:")
            dpg.add_input_text(tag="anim_title_input", default_value="Test", width=-1)
        with dpg.group(horizontal=True):
            dpg.add_text("Merge:")
            dpg.add_input_int(tag="anim_merge_input", default_value=1, min_value=0,
                              max_value=3, width=-1, step=0)
        with dpg.group(horizontal=True):
            dpg.add_text("Max NFEs:")
            dpg.add_input_int(tag="anim_nfes_input", default_value=1000, min_value=1,
                              width=-1, step=0)
        with dpg.group(horizontal=True):
            dpg.add_text("Format:")
            dpg.add_combo(["gif", "mp4"], tag="anim_format_combo", default_value="gif", width=-1)
        dpg.add_checkbox(label="Log Scale", tag="anim_logscale_check", default_value=False)
        dpg.add_spacer(height=8)
        dpg.add_separator()
        dpg.add_spacer(height=5)
        with dpg.group(horizontal=True):
            dpg.add_button(label="Generate Animation", width=160, callback=_run_animation_generation)
            dpg.add_spacer(width=10)
            dpg.add_button(label="Cancel", width=80,
                           callback=lambda: dpg.delete_item(modal_tag))


def _run_animation_generation(sender, app_data):
    """Run animation generation in a background thread."""
    title = dpg.get_value("anim_title_input") or "Test"
    merge = dpg.get_value("anim_merge_input")
    max_nfes = dpg.get_value("anim_nfes_input") or 1000
    fmt = dpg.get_value("anim_format_combo") or "gif"
    log_scale = dpg.get_value("anim_logscale_check")

    # Close the dialog
    if dpg.does_item_exist("animation_modal"):
        dpg.delete_item("animation_modal")

    # Determine data and results paths
    data_path = _state["file_manager"].get_data_path_str()
    results_path = _state["file_manager"].get_results_path_str()

    # Show progress in results area
    if dpg.does_item_exist("test_results_area"):
        dpg.delete_item("test_results_area", children_only=True)
        dpg.add_text("Generating animation...", parent="test_results_area", color=COLOR_TITLE)

    def _do_generate():
        try:
            from ddmtolab.Methods.animation_generator import AnimationGenerator
            generator = AnimationGenerator(
                data_path=data_path,
                save_path=results_path,
                title=title,
                merge=merge,
                max_nfes=max_nfes,
                format=fmt,
                log_scale=log_scale,
            )
            result = generator.run()
            _state["_anim_result"] = result
            _state["_anim_error"] = None
        except Exception as e:
            _state["_anim_result"] = None
            _state["_anim_error"] = str(e)

    _state["_anim_result"] = None
    _state["_anim_error"] = None
    _state["_anim_thread"] = threading.Thread(target=_do_generate, daemon=True)
    _state["_anim_thread"].start()


def _display_loaded_results(results_path: Path):
    """Display analysis outputs from a loaded data folder."""
    if not results_path.exists():
        dpg.add_text("Results folder not found.", parent="test_results_area", color=(180, 80, 80))
        return

    viewport_w = dpg.get_viewport_width()
    available_width = max(400, viewport_w - 250 - 280 - 50)
    IMG_WIDTH = (available_width - 40) // 3
    per_row = 3

    # Open folder button
    def _open_folder():
        import subprocess
        import sys
        if sys.platform == 'win32':
            os.startfile(str(results_path.resolve()))
        elif sys.platform == 'darwin':
            subprocess.run(['open', str(results_path.resolve())])
        else:
            subprocess.run(['xdg-open', str(results_path.resolve())])

    with dpg.group(horizontal=True, parent="test_results_area"):
        dpg.add_spacer(width=-1)
        btn = dpg.add_button(label="Open Results Folder", callback=lambda: _open_folder(), width=150)
        with dpg.theme() as btn_theme:
            with dpg.theme_component(dpg.mvButton):
                dpg.add_theme_color(dpg.mvThemeCol_Text, (40, 40, 40))
                dpg.add_theme_color(dpg.mvThemeCol_Button, (200, 200, 200))
                dpg.add_theme_color(dpg.mvThemeCol_ButtonHovered, (180, 180, 180))
                dpg.add_theme_color(dpg.mvThemeCol_ButtonActive, (160, 160, 160))
        dpg.bind_item_theme(btn, btn_theme)

    dpg.add_spacer(height=5, parent="test_results_area")

    # Excel tables
    excel_files = list(results_path.glob("*.xlsx"))
    if excel_files:
        dpg.add_text("Results Table", parent="test_results_area", color=COLOR_TITLE)
        dpg.add_separator(parent="test_results_area")
        with dpg.theme() as table_theme:
            with dpg.theme_component(dpg.mvTable):
                dpg.add_theme_color(dpg.mvThemeCol_TableHeaderBg, (70, 70, 70))
                dpg.add_theme_color(dpg.mvThemeCol_Text, (255, 255, 255))
            with dpg.theme_component(dpg.mvTableRow):
                dpg.add_theme_color(dpg.mvThemeCol_Text, (40, 40, 40))
        for excel_file in excel_files:
            try:
                import pandas as pd
                df = pd.read_excel(excel_file)
                df = df.loc[:, ~df.columns.str.contains('^Unnamed')]
                df = df.dropna(axis=1, how='all')
                cols = list(df.columns)
                tbl = dpg.add_table(parent="test_results_area", header_row=True,
                                    borders_innerH=True, borders_outerH=True,
                                    borders_innerV=True, borders_outerV=True,
                                    policy=dpg.mvTable_SizingStretchSame)
                dpg.bind_item_theme(tbl, table_theme)
                for i, col in enumerate(cols):
                    col_width = 120 if i == 0 else 140
                    dpg.add_table_column(label=str(col), parent=tbl, width_fixed=True,
                                         init_width_or_weight=col_width)
                for _, row in df.iterrows():
                    with dpg.table_row(parent=tbl):
                        for col in cols:
                            dpg.add_text(str(row[col]))
            except Exception as e:
                dpg.add_text(f"Could not load table: {e}", parent="test_results_area",
                             color=(200, 60, 60))

    # PNG images
    all_images = sorted(results_path.glob("*.png"))
    convergence_images = [f for f in all_images if "convergence" in f.name.lower()]
    runtime_images = [f for f in all_images if "runtime" in f.name.lower()]
    other_images = [f for f in all_images if f not in convergence_images and f not in runtime_images]
    nd_folder = results_path / "ND_Solutions"
    nd_images = sorted(nd_folder.glob("*.png")) if nd_folder.exists() else []

    for section_name, images in [("Convergence Curves", convergence_images),
                                  ("Non-Dominated Solutions", nd_images),
                                  ("Runtime Comparison", runtime_images),
                                  ("Other", other_images)]:
        if images:
            dpg.add_spacer(height=10, parent="test_results_area")
            dpg.add_text(section_name, parent="test_results_area", color=COLOR_TITLE)
            dpg.add_separator(parent="test_results_area")
            for i in range(0, len(images), per_row):
                row_files = images[i:i + per_row]
                with dpg.group(horizontal=True, parent="test_results_area"):
                    for img_path in row_files:
                        _load_and_display_image(str(img_path), 0, max_width=IMG_WIDTH)
                        dpg.add_spacer(width=10)
                dpg.add_spacer(height=8, parent="test_results_area")


def _load_data_clicked(sender, app_data):
    """Open folder dialog to load existing data and run analysis."""
    import tkinter as tk
    from tkinter import filedialog

    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)

    # Open parent folder so Data folder is visible and selectable
    default_path = str(Path(_state["file_manager"].get_data_path_str()).parent)

    folder_path = filedialog.askdirectory(
        title="Select Data Folder",
        initialdir=default_path,
    )

    root.destroy()

    if not folder_path:
        return

    data_dir = Path(folder_path)
    pkl_files = list(data_dir.rglob("*.pkl"))
    if not pkl_files:
        show_error_modal("No .pkl data files found in selected folder.")
        return

    save_path = str(data_dir.parent / "Results")

    # Read analysis settings from UI
    figure_format = dpg.get_value("test_fig_combo") if dpg.does_item_exist("test_fig_combo") else "png"
    use_log = dpg.get_value("test_log_check") if dpg.does_item_exist("test_log_check") else False
    show_pf = dpg.get_value("test_showpf_check") if dpg.does_item_exist("test_showpf_check") else True
    show_nd = dpg.get_value("test_shownd_check") if dpg.does_item_exist("test_shownd_check") else True
    best_so_far = dpg.get_value("test_bestsofar_check") if dpg.does_item_exist("test_bestsofar_check") else True

    # Lock run button
    _state["running"] = True
    dpg.configure_item("test_run_btn", enabled=False)
    dpg.bind_item_theme("test_run_btn", _get_disabled_btn_theme())

    dpg.delete_item("test_results_area", children_only=True)
    dpg.add_text("Analyzing loaded data...", parent="test_results_area",
                 tag="test_running_text", color=(60, 60, 60))
    dpg.add_progress_bar(tag="test_progress", parent="test_results_area",
                         default_value=0.5, width=-1)
    dpg.add_text("", parent="test_results_area",
                 tag="test_elapsed_text", color=(100, 100, 100))

    def _worker():
        start_time = time.time()
        try:
            from ddmtolab.Methods.test_data_analysis import TestDataAnalyzer
            analyzer = TestDataAnalyzer(
                data_path=folder_path,
                save_path=save_path,
                figure_format=figure_format,
                log_scale=use_log,
                show_pf=show_pf,
                show_nd=show_nd,
                best_so_far=best_so_far,
                clear_results=True,
            )
            analyzer.run()
        except Exception as e:
            import traceback
            _state["statuses"]["_load"] = type('S', (), {
                'error': f"{e}\n{traceback.format_exc()}", 'finished': True,
                'running': False, 'result': None, 'elapsed': 0})()
        finally:
            _state["running"] = False
            # Store save path for display
            _state["_load_save_path"] = save_path
            _state["_load_elapsed"] = time.time() - start_time

    _state["statuses"] = {}
    _state["results"] = {}
    _state["results_displayed"] = False
    _state["_load_save_path"] = None

    t = threading.Thread(target=_worker, daemon=True)
    t.start()


def create(parent, base_path: str = "./tests"):
    """Create the Test Mode UI inside the given parent container."""
    _init_managers(base_path)

    with dpg.group(horizontal=True, parent=parent):
        # Left panel - Problem Selection (Column 1)
        with dpg.child_window(width=250, tag="test_problem_panel"):
            dpg.add_text("Problem Selection", color=COLOR_TITLE)
            dpg.add_separator()

            dpg.add_text("Category", color=COLOR_SUCCESS)
            dpg.add_combo(CATEGORIES, default_value="STSO", tag="test_prob_cat_combo",
                          callback=_on_prob_category_change, width=-1)

            dpg.add_spacer(height=5)
            dpg.add_text("Suite", color=COLOR_SUCCESS)
            initial_suites = get_problem_suites("STSO")
            dpg.add_combo(initial_suites, default_value=initial_suites[0] if initial_suites else "",
                          tag="test_suite_combo", callback=_on_suite_change, width=-1)

            dpg.add_spacer(height=5)
            dpg.add_text("Problem", color=COLOR_SUCCESS)
            initial_methods = get_problem_methods("STSO", initial_suites[0]) if initial_suites else []
            dpg.add_combo(initial_methods, default_value=initial_methods[0] if initial_methods else "",
                          tag="test_method_combo", width=-1)

            # Problem Parameters
            dpg.add_spacer(height=10)
            with dpg.group(tag="test_problem_params_group"):
                dpg.add_text("Problem Parameters", color=COLOR_TITLE)
                dpg.add_separator()

                with dpg.group(tag="test_D_group", horizontal=True):
                    dpg.add_text("D:")
                    dpg.add_input_int(default_value=50, min_value=2, max_value=100,
                                      tag="test_D_input", width=-1, step=0)

                with dpg.group(tag="test_M_group", horizontal=True, show=False):
                    dpg.add_text("M:")
                    dpg.add_input_int(default_value=3, min_value=2, max_value=10,
                                      tag="test_M_input", width=-1, step=0)

                with dpg.group(tag="test_K_group", horizontal=True, show=False):
                    dpg.add_text("K:")
                    dpg.add_input_int(default_value=10, min_value=2, max_value=100,
                                      tag="test_K_input", width=-1, step=0)

                with dpg.group(tag="test_L_group", horizontal=True, show=False):
                    dpg.add_text("L:")
                    dpg.add_input_int(default_value=20, min_value=1, max_value=50,
                                      tag="test_L_input", width=-1, step=0)

            dpg.add_spacer(height=10)

        # Middle panel - Algorithm Selection (Column 2)
        with dpg.child_window(width=280, tag="test_algo_panel"):
            dpg.add_text("Algorithm Selection", color=COLOR_TITLE)
            dpg.add_separator()

            dpg.add_text("Category", color=COLOR_SUCCESS)
            dpg.add_combo(CATEGORIES, default_value="STSO", tag="test_algo_cat_combo",
                          callback=_on_algo_category_change, width=-1)

            dpg.add_spacer(height=5)
            with dpg.child_window(tag="test_algo_list_container", height=215, border=True):
                # Will be populated by _update_algo_list_display
                pass

            # Selected Algorithms with parameters (dynamic)
            dpg.add_spacer(height=10)
            dpg.add_text("Selected Algorithms", color=COLOR_TITLE)
            dpg.add_separator()
            with dpg.child_window(tag="test_algo_params_container", height=-1, border=False):
                pass

        # Right panel - Options and Results
        with dpg.child_window(tag="test_right_panel"):
            # Buttons row at the top
            with dpg.group(horizontal=True):
                dpg.add_button(label="Load Data", callback=_load_data_clicked, width=90)
                dpg.add_spacer(width=8)
                dpg.add_button(label="Clean Data", callback=_clean_clicked, width=80)
                dpg.add_spacer(width=8)
                run_btn = dpg.add_button(label="Run", tag="test_run_btn",
                                         callback=_run_clicked, width=100)
                dpg.add_spacer(width=8)
                stop_btn = dpg.add_button(label="Stop", tag="test_stop_btn",
                                          callback=_stop_clicked, width=100)
                dpg.add_spacer(width=8)
                anim_btn = dpg.add_button(label="Animation", callback=_animation_clicked, width=100)

            # Apply colored themes to Run and Stop buttons
            with dpg.theme() as run_theme:
                with dpg.theme_component(dpg.mvButton):
                    dpg.add_theme_color(dpg.mvThemeCol_Button, (60, 140, 60))
                    dpg.add_theme_color(dpg.mvThemeCol_ButtonHovered, (80, 160, 80))
                    dpg.add_theme_color(dpg.mvThemeCol_ButtonActive, (100, 180, 100))
            dpg.bind_item_theme(run_btn, run_theme)
            _state["run_theme"] = run_theme

            with dpg.theme() as stop_theme:
                with dpg.theme_component(dpg.mvButton):
                    dpg.add_theme_color(dpg.mvThemeCol_Button, (160, 60, 60))
                    dpg.add_theme_color(dpg.mvThemeCol_ButtonHovered, (180, 80, 80))
                    dpg.add_theme_color(dpg.mvThemeCol_ButtonActive, (200, 100, 100))
            dpg.bind_item_theme(stop_btn, stop_theme)

            with dpg.theme() as delete_theme:
                with dpg.theme_component(dpg.mvButton):
                    dpg.add_theme_color(dpg.mvThemeCol_Button, (140, 50, 50))
                    dpg.add_theme_color(dpg.mvThemeCol_ButtonHovered, (170, 70, 70))
                    dpg.add_theme_color(dpg.mvThemeCol_ButtonActive, (200, 90, 90))
            _state["delete_theme"] = delete_theme

            with dpg.theme() as anim_theme:
                with dpg.theme_component(dpg.mvButton):
                    dpg.add_theme_color(dpg.mvThemeCol_Button, (180, 120, 40))
                    dpg.add_theme_color(dpg.mvThemeCol_ButtonHovered, (200, 140, 60))
                    dpg.add_theme_color(dpg.mvThemeCol_ButtonActive, (220, 160, 80))
            dpg.bind_item_theme(anim_btn, anim_theme)

            # Apply bold font to Run and Stop buttons
            from main import get_fonts
            _bold = get_fonts().get("bold")
            if _bold:
                dpg.bind_item_font(run_btn, _bold)
                dpg.bind_item_font(stop_btn, _bold)

            dpg.add_spacer(height=5)
            dpg.add_separator()

            # Analysis Settings (same layout as Batch Experiment)
            dpg.add_text("Analysis Settings", color=COLOR_TITLE)
            dpg.add_separator()

            # Row 1: Pure checkboxes
            with dpg.group(horizontal=True):
                dpg.add_checkbox(label="Log Scale", tag="test_log_check",
                                 default_value=True)
                dpg.add_spacer(width=5)
                dpg.add_checkbox(label="Best-So-Far", tag="test_bestsofar_check",
                                 default_value=True)
                dpg.add_spacer(width=5)
                dpg.add_checkbox(label="Show ND", tag="test_shownd_check",
                                 default_value=True)
                dpg.add_spacer(width=5)
                dpg.add_checkbox(label="Show PF", tag="test_showpf_check",
                                 default_value=True)
                dpg.add_spacer(width=15)
                dpg.add_checkbox(label="Std Band", tag="test_stdband_check",
                                 default_value=False)
                dpg.add_spacer(width=5)
                dpg.add_checkbox(label="Merge Plots", tag="test_merge_check",
                                 default_value=False)

            dpg.add_spacer(height=3)

            # Row 2: Combos and input parameters
            with dpg.group(horizontal=True):
                dpg.add_text("Metric")
                dpg.add_combo(METRICS, default_value="IGD",
                              tag="test_metric_combo", width=80)
                dpg.add_spacer(width=5)
                dpg.add_text("Figure")
                dpg.add_combo(FIGURE_FORMATS, default_value="png",
                              tag="test_fig_combo", width=70)

            dpg.add_spacer(height=5)
            dpg.add_separator()

            # Results area (white background for seamless image display)
            with dpg.child_window(tag="test_results_area") as results_panel:
                dpg.add_text("No results yet. Configure and run an experiment.", color=(100, 100, 100))

        # Apply white background theme to results area with dark text
        with dpg.theme() as results_theme:
            with dpg.theme_component(dpg.mvAll):
                dpg.add_theme_color(dpg.mvThemeCol_ChildBg, (255, 255, 255))
                dpg.add_theme_color(dpg.mvThemeCol_Text, (40, 40, 40))
                dpg.add_theme_color(dpg.mvThemeCol_FrameBg, (245, 245, 245))
                dpg.add_theme_color(dpg.mvThemeCol_Border, (200, 200, 200))
        dpg.bind_item_theme(results_panel, results_theme)

    # Initialize algorithm list
    _update_algo_list_display()


def update():
    """Called each frame to check if results are ready."""
    # Check if animation generation finished
    anim_thread = _state.get("_anim_thread")
    if anim_thread is not None and not anim_thread.is_alive():
        _state["_anim_thread"] = None
        if dpg.does_item_exist("test_results_area"):
            dpg.delete_item("test_results_area", children_only=True)
        error = _state.get("_anim_error")
        if error:
            dpg.add_text(f"Animation failed: {error}", parent="test_results_area", color=COLOR_ERROR)
        else:
            result = _state.get("_anim_result", {})
            success = result.get("success", [])
            failed = result.get("failed", [])
            results_path = _state["file_manager"].get_results_path_str()
            dpg.add_text(f"Animation completed! Success: {len(success)}, Failed: {len(failed)}",
                         parent="test_results_area", color=COLOR_SUCCESS)
            dpg.add_text(f"Saved to: {results_path}", parent="test_results_area")

    # Check if load-data analysis finished
    if not _state["running"] and _state.get("_load_save_path") and not _state.get("results_displayed", False):
        _state["results_displayed"] = True
        dpg.configure_item("test_run_btn", enabled=True)
        if "run_theme" in _state:
            dpg.bind_item_theme("test_run_btn", _state["run_theme"])
        # Display loaded results
        save_path = _state["_load_save_path"]
        _state["_load_save_path"] = None
        dpg.delete_item("test_results_area", children_only=True)
        elapsed = _state.get("_load_elapsed", 0)
        # Check for errors
        load_status = _state.get("statuses", {}).get("_load")
        if load_status and load_status.error:
            error_msg = f"Analysis failed: {load_status.error}"
            with dpg.group(horizontal=True, parent="test_results_area"):
                dpg.add_text(error_msg, color=(200, 60, 60), wrap=760)
                dpg.add_button(label="\u2398", callback=_on_copy_error_click,
                               user_data=error_msg, width=24)
        else:
            dpg.add_text(f"Analysis completed in {format_time(elapsed)}",
                         parent="test_results_area", color=(60, 160, 60))
            # Reuse _display_results logic by pointing results_path to load save_path
            results_path = Path(save_path)
            _display_loaded_results(results_path)
        return

    # Check if experiment finished (results exist and not running)
    if not _state["running"] and _state.get("statuses") and not _state.get("results_displayed", False):
        _state["results_displayed"] = True
        # Unlock run button
        dpg.configure_item("test_run_btn", enabled=True)
        if "run_theme" in _state:
            dpg.bind_item_theme("test_run_btn", _state["run_theme"])
        _display_results()
