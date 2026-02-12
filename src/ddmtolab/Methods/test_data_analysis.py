"""
Test Data Analyzer Module for Single-Run Algorithm Testing

This module provides a lightweight analysis pipeline for quick algorithm testing,
reading pickle files directly from the data folder for single-run visualization.

Classes:
    TestScanResult: Dataclass for storing test scan results
    TestMetricResults: Dataclass for storing test metric results
    TestDataAnalyzer: Main class for test data analysis

Usage:
    analyzer = TestDataAnalyzer(data_path='./Data', save_path='./Results')
    results = analyzer.run()

Author: Jiangtao Shen
Email: j.shen5@exeter.ac.uk
Date: 2025.10.10
Version: 2.0
"""

import shutil
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple, Union
from dataclasses import dataclass, field
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.ticker import ScalarFormatter, LogFormatterSciNotation

# Import from project modules
from ddmtolab.Methods.metrics import IGD, HV, GD, IGDp, FR, CV, DeltaP, Spread, Spacing
from ddmtolab.Methods.Algo_Methods.algo_utils import nd_sort
from ddmtolab.Methods.data_analysis import (
    OptimizationDirection,
    DataUtils,
    DEFAULT_COLORS,
    DEFAULT_MARKERS,
)


# =============================================================================
# Data Classes
# =============================================================================

@dataclass
class TestScanResult:
    """
    Result of scanning a test data directory.

    Attributes:
        algorithms: List[str]
            Sorted list of algorithm names found in the directory.
        problems: List[str]
            List of problem names (extracted from data or default).
        data_path: Path
            Path to the scanned data directory.
        file_mapping: Dict[str, Path]
            Mapping from algorithm name to pickle file path.
    """
    algorithms: List[str]
    problems: List[str]
    data_path: Path
    file_mapping: Dict[str, Path]


@dataclass
class TestMetricResults:
    """
    Container for test metric calculation results.

    Attributes:
        metric_values: Dict[str, List[np.ndarray]]
            Metric values per generation per algorithm.
            Structure: metric_values[algorithm] = List[np.ndarray] (per task)

        best_values: Dict[str, List[float]]
            Final best metric values per algorithm.
            Structure: best_values[algorithm] = List[float] (per task)

        objective_values: Dict[str, List[np.ndarray]]
            Original objective values per algorithm.
            Structure: objective_values[algorithm] = List[np.ndarray]

        runtime: Dict[str, float]
            Runtime in seconds per algorithm.

        max_nfes: Dict[str, List[int]]
            Maximum number of function evaluations per algorithm.

        metric_name: Optional[str]
            Name of the metric used.

        problems: List[str]
            List of problem names.
    """
    metric_values: Dict[str, List[np.ndarray]]
    best_values: Dict[str, List[float]]
    objective_values: Dict[str, List[np.ndarray]]
    runtime: Dict[str, float]
    max_nfes: Dict[str, List[int]]
    metric_name: Optional[str]
    problems: List[str]


@dataclass
class PlotConfig:
    """
    Configuration for plot generation.

    Attributes:
        figure_format: str
            Output figure format (e.g., 'pdf', 'png', 'svg').
        log_scale: bool
            Whether to use logarithmic scale for y-axis.
        show_pf: bool
            Whether to show true Pareto front in ND solution plots.
        show_nd: bool
            Whether to filter and show only non-dominated solutions.
        save_path: Path
            Directory path to save output figures.
        colors: List[str]
            Color palette for plotting algorithms.
        markers: List[str]
            Marker styles for plotting algorithms.
    """
    figure_format: str = 'pdf'
    log_scale: bool = False
    show_pf: bool = True
    show_nd: bool = True
    save_path: Path = Path('./Results')
    colors: List[str] = field(default_factory=lambda: DEFAULT_COLORS.copy())
    markers: List[str] = field(default_factory=lambda: DEFAULT_MARKERS.copy())


# =============================================================================
# Test Plot Generator
# =============================================================================

class TestPlotGenerator:
    """Class for generating visualization plots for test data."""

    def __init__(self, config: PlotConfig):
        """Initialize TestPlotGenerator with configuration."""
        self.config = config

    @staticmethod
    def _calculate_legend_fontsize(n_algorithms: int) -> int:
        """
        Calculate legend font size based on number of algorithms.

        Linear interpolation:
        - 2 algorithms -> font size 14
        - 15 algorithms -> font size 6
        """
        if n_algorithms <= 2:
            return 14
        elif n_algorithms >= 15:
            return 6
        else:
            return int(round(14 - (8 / 13) * (n_algorithms - 2)))

    def plot_convergence_curves(
            self,
            metric_values: Dict[str, List[np.ndarray]],
            max_nfes: Dict[str, List[int]],
            algorithm_order: List[str],
            problems: List[str],
            metric_name: Optional[str] = None
    ) -> None:
        """Generate and save convergence curve plots for all algorithms and tasks."""
        save_dir = Path(self.config.save_path)
        save_dir.mkdir(parents=True, exist_ok=True)

        first_algo = algorithm_order[0]
        num_tasks = len(metric_values[first_algo])

        for task_idx in range(num_tasks):
            fig = self._create_convergence_figure(
                metric_values, max_nfes, algorithm_order,
                task_idx, num_tasks, problems, metric_name
            )

            if num_tasks == 1:
                problem_name = problems[0] if problems else 'Test'
                output_file = save_dir / f'{problem_name}_convergence.{self.config.figure_format}'
            else:
                problem_name = problems[task_idx] if task_idx < len(problems) else f'P{task_idx + 1}'
                output_file = save_dir / f'{problem_name}-Task{task_idx + 1}_convergence.{self.config.figure_format}'

            fig.savefig(output_file, dpi=300, bbox_inches='tight')
            plt.close(fig)

        print(f"Convergence plots saved to: {save_dir}")

    def _create_convergence_figure(
            self,
            metric_values: Dict[str, List[np.ndarray]],
            max_nfes: Dict[str, List[int]],
            algorithm_order: List[str],
            task_idx: int,
            num_tasks: int,
            problems: List[str],
            metric_name: Optional[str]
    ) -> plt.Figure:
        """Create a single convergence curve figure."""
        fig, ax = plt.subplots(figsize=(5, 3.5))

        # Collect curve data for y-axis range and max NFEs for x-axis formatting
        all_curves = []
        actual_max_nfes = 0

        # Adaptive line width and marker size based on number of algorithms
        n_algos = len(algorithm_order)
        if n_algos <= 4:
            markersize, linewidth = 8, 2.5
        elif n_algos <= 6:
            markersize, linewidth = 7, 2.0
        else:
            markersize, linewidth = 6, 1.6

        for idx, algo in enumerate(algorithm_order):
            curve = np.array(metric_values[algo][task_idx]).ravel()

            if len(curve) == 0:
                continue

            all_curves.append(curve)

            nfes = max_nfes[algo][task_idx] if task_idx < len(max_nfes[algo]) else len(curve)
            actual_max_nfes = max(actual_max_nfes, nfes)
            x = np.linspace(0, nfes, len(curve))
            marker_interval = max(1, len(curve) // 10)

            ax.plot(
                x, curve, label=algo,
                color=self.config.colors[idx % len(self.config.colors)],
                marker=self.config.markers[idx % len(self.config.markers)],
                markevery=marker_interval,
                markersize=markersize, linewidth=linewidth, linestyle='-', alpha=0.7
            )

        # Set axis labels
        y_label = metric_name if metric_name is not None else 'Objective Value'
        ax.set_xlabel('NFEs', fontsize=14)
        ax.set_ylabel(y_label, fontsize=14)

        if num_tasks == 1:
            title = problems[0] if problems else 'Test Problem'
        else:
            prob_name = problems[task_idx] if task_idx < len(problems) else f'P{task_idx + 1}'
            title = f'{prob_name} - Task {task_idx + 1}'
        ax.set_title(title, fontsize=14)
        ax.tick_params(axis='both', which='major', labelsize=14)

        # Auto-adjust legend font size based on number of algorithms
        legend_fontsize = self._calculate_legend_fontsize(len(algorithm_order))
        ax.legend(loc='best', fontsize=legend_fontsize)
        ax.grid(True, alpha=0.2, linestyle='-')

        # Apply axis formatting
        if self.config.log_scale:
            ax.set_yscale('log')
            # Check data range; use linear scale if range is too small
            if len(all_curves) > 0:
                all_data = np.concatenate([c for c in all_curves])
                y_min, y_max = np.min(all_data), np.max(all_data)

                if y_max / y_min < 10:
                    ax.set_yscale('linear')
                    self._apply_scientific_notation(ax, actual_xmax=actual_max_nfes)
                else:
                    ax.yaxis.set_major_formatter(LogFormatterSciNotation())
                    if actual_max_nfes > 10000:
                        formatter = ScalarFormatter(useMathText=True)
                        formatter.set_scientific(True)
                        formatter.set_powerlimits((0, 0))
                        ax.xaxis.set_major_formatter(formatter)
        else:
            self._apply_scientific_notation(ax, actual_xmax=actual_max_nfes)

        # Disable minor ticks (must be called after set_yscale)
        ax.minorticks_off()

        fig.tight_layout()
        return fig

    def _apply_scientific_notation(
            self,
            ax: plt.Axes,
            actual_xmax: Optional[float] = None,
            x_threshold: float = 10000,
            y_threshold: float = 1000
    ) -> None:
        """Apply scientific notation to axes if values exceed threshold."""
        xmax = actual_xmax if actual_xmax is not None else ax.get_xlim()[1]
        ymax = ax.get_ylim()[1]

        if xmax > x_threshold:
            formatter = ScalarFormatter(useMathText=True)
            formatter.set_scientific(True)
            formatter.set_powerlimits((0, 0))
            ax.xaxis.set_major_formatter(formatter)

        if ymax > y_threshold:
            formatter = ScalarFormatter(useMathText=True)
            formatter.set_scientific(True)
            formatter.set_powerlimits((0, 0))
            ax.yaxis.set_major_formatter(formatter)

    def plot_runtime(
            self,
            runtime: Dict[str, float],
            algorithm_order: List[str]
    ) -> None:
        """Generate and save a bar plot showing runtime comparison."""
        save_dir = Path(self.config.save_path)
        save_dir.mkdir(parents=True, exist_ok=True)

        n_algos = len(algorithm_order)

        # Adaptive figure width based on number of algorithms
        fig_width = max(4, min(6, 2 + n_algos * 0.8))
        fig, ax = plt.subplots(figsize=(fig_width, 3.5))

        x = np.arange(n_algos)
        runtimes = [runtime[algo] for algo in algorithm_order]
        colors = [self.config.colors[i % len(self.config.colors)] for i in range(n_algos)]

        # Adaptive bar width: narrower bars when fewer algorithms
        if n_algos == 1:
            bar_width = 0.35
        elif n_algos == 2:
            bar_width = 0.45
        elif n_algos <= 4:
            bar_width = 0.55
        else:
            bar_width = 0.7

        bars = ax.bar(x, runtimes, width=bar_width, color=colors, alpha=0.8,
                      edgecolor='black', linewidth=0.8)

        # Add value labels on bars
        for bar, val in zip(bars, runtimes):
            height = bar.get_height()
            ax.annotate(f'{val:.2f}s',
                        xy=(bar.get_x() + bar.get_width() / 2, height),
                        xytext=(0, 3),
                        textcoords="offset points",
                        ha='center', va='bottom', fontsize=10)

        ax.set_ylabel('Runtime (s)', fontsize=12)
        ax.set_xticks(x)
        ax.set_xticklabels(algorithm_order, fontsize=12)
        ax.tick_params(axis='both', which='major', labelsize=10)
        ax.grid(True, axis='y', alpha=0.3, linestyle='-')

        # Add top margin for labels (15% extra space)
        y_max = max(runtimes) if runtimes else 1
        ax.set_ylim(0, y_max * 1.15)

        fig.tight_layout()

        output_file = save_dir / f'runtime_comparison.{self.config.figure_format}'
        fig.savefig(output_file, dpi=300, bbox_inches='tight')
        plt.close(fig)

        print(f"Runtime plot saved to: {output_file}")

    def plot_nd_solutions(
            self,
            objective_values: Dict[str, List[np.ndarray]],
            algorithm_order: List[str],
            problems: List[str],
            settings: Optional[Dict[str, Any]] = None
    ) -> None:
        """Generate and save non-dominated solution plots."""
        nd_folder = Path(self.config.save_path) / 'ND_Solutions'
        nd_folder.mkdir(parents=True, exist_ok=True)

        first_algo = algorithm_order[0]
        n_tasks = len(objective_values[first_algo])

        for algo in algorithm_order:
            for task_idx in range(n_tasks):
                objectives = objective_values[algo][task_idx]

                if objectives is None or objectives.shape[0] == 0:
                    continue

                n_objectives = objectives.shape[1]

                if n_objectives <= 1:
                    continue

                # Filter non-dominated solutions if requested
                if self.config.show_nd:
                    front_no, _ = nd_sort(objectives, objectives.shape[0])
                    nd_solutions = objectives[front_no == 1]
                else:
                    nd_solutions = objectives

                # Load true Pareto front if requested
                true_pf = None
                if self.config.show_pf and settings is not None:
                    prob = problems[task_idx] if task_idx < len(problems) else f'P{task_idx + 1}'
                    true_pf = DataUtils.load_reference(settings, prob, task_idx, n_objectives)

                # Create plot
                fig = self._create_nd_plot(nd_solutions, true_pf, n_objectives, n_tasks,
                                           problems, task_idx, algo)

                # Save figure
                prob_name = problems[task_idx] if task_idx < len(problems) else f'P{task_idx + 1}'
                if n_tasks == 1:
                    filename = f'{prob_name}-{algo}.{self.config.figure_format}'
                else:
                    filename = f'{prob_name}-Task{task_idx + 1}-{algo}.{self.config.figure_format}'
                fig.savefig(nd_folder / filename, dpi=300)
                plt.close(fig)

        print(f"ND solution plots saved to: {nd_folder}")

    def _create_nd_plot(
            self,
            nd_solutions: np.ndarray,
            true_pf: Optional[np.ndarray],
            n_objectives: int,
            n_tasks: int,
            problems: List[str],
            task_idx: int,
            algo: str
    ) -> plt.Figure:
        """Create a non-dominated solution plot."""
        fig = plt.figure(figsize=(4.5, 3.5))

        if n_objectives == 2:
            ax = fig.add_subplot(111)

            if true_pf is not None and true_pf.shape[1] == 2:
                sort_idx = np.argsort(true_pf[:, 0])
                sorted_pf = true_pf[sort_idx]
                ax.scatter(sorted_pf[:, 0], sorted_pf[:, 1],
                           c='gray', s=2, linewidth=0.1, zorder=1)

            ax.scatter(nd_solutions[:, 0], nd_solutions[:, 1],
                       c='dodgerblue', s=60, alpha=0.8, edgecolors='black',
                       linewidth=0.8, zorder=2)

            ax.set_xlabel('$f_1$', fontsize=12)
            ax.set_ylabel('$f_2$', fontsize=12)
            ax.grid(True, alpha=0.2, linestyle='-')

        elif n_objectives == 3:
            ax = fig.add_subplot(111, projection='3d')

            if true_pf is not None and true_pf.shape[1] == 3:
                ax.scatter(true_pf[:, 0], true_pf[:, 1], true_pf[:, 2],
                           c='gray', s=4, alpha=0.2, zorder=1, depthshade=True)

            ax.scatter(nd_solutions[:, 0], nd_solutions[:, 1], nd_solutions[:, 2],
                       c='dodgerblue', s=60, alpha=0.8, edgecolors='black',
                       linewidth=0.8, zorder=2, depthshade=True)

            ax.set_xlabel('$f_1$', fontsize=12)
            ax.set_ylabel('$f_2$', fontsize=12)
            ax.set_zlabel('$f_3$', fontsize=12)
            ax.view_init(elev=20, azim=60)

        else:
            # Parallel coordinates for many-objective
            ax = fig.add_subplot(111)

            for i in range(nd_solutions.shape[0]):
                ax.plot(range(n_objectives), nd_solutions[i, :],
                        'b-', alpha=0.3, linewidth=0.8)

            ax.set_xlabel('Objective', fontsize=12)
            ax.set_ylabel('Value', fontsize=12)
            ax.set_xticks(range(n_objectives))
            ax.set_xticklabels([rf'$f_{{{i + 1}}}$' for i in range(n_objectives)])
            ax.grid(True, alpha=0.3, linestyle='--')

        prob_name = problems[task_idx] if task_idx < len(problems) else f'P{task_idx + 1}'
        if n_tasks == 1:
            title = f'{prob_name} - {algo}'
        else:
            title = f'{prob_name} - Task {task_idx + 1} - {algo}'
        plt.title(title, fontsize=10)
        plt.tight_layout()

        return fig


# =============================================================================
# Test Table Generator
# =============================================================================

class TestTableGenerator:
    """Class for generating Excel tables for test data."""

    def __init__(self, save_path: Path = Path('./Results')):
        """Initialize TestTableGenerator."""
        self.save_path = save_path

    def generate_excel_table(
            self,
            best_values: Dict[str, List[float]],
            runtime: Dict[str, float],
            algorithm_order: List[str],
            problems: List[str],
            metric_name: Optional[str] = None
    ) -> Path:
        """
        Generate an Excel table comparing algorithm performance.

        Parameters:
            best_values: Dict[str, List[float]]
                Final metric values per algorithm per task.
            runtime: Dict[str, float]
                Runtime per algorithm.
            algorithm_order: List[str]
                Algorithm display order.
            problems: List[str]
                Problem/task names.
            metric_name: Optional[str]
                Metric name for table caption.

        Returns:
            Path: Path to generated Excel file.
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            print("Warning: openpyxl not installed. Skipping Excel table generation.")
            return None

        save_dir = Path(self.save_path)
        save_dir.mkdir(parents=True, exist_ok=True)

        num_tasks = len(best_values[algorithm_order[0]])
        direction = DataUtils.get_metric_direction(metric_name)

        # Create workbook
        wb = Workbook()
        ws = wb.active
        metric_str = metric_name if metric_name else 'Objective'
        ws.title = f"Test Results ({metric_str})"

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        best_font = Font(bold=True)
        center_align = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Header row
        headers = ['Problem'] + algorithm_order
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        # Data rows
        row_idx = 2
        for task_idx in range(num_tasks):
            prob_name = problems[task_idx] if task_idx < len(problems) else f'Task {task_idx + 1}'

            # Find best value for this task
            values = {}
            for algo in algorithm_order:
                val = best_values[algo][task_idx]
                values[algo] = val

            valid_values = {k: v for k, v in values.items() if not np.isnan(v)}
            best_algo = None
            if valid_values:
                if direction == OptimizationDirection.MINIMIZE:
                    best_algo = min(valid_values, key=valid_values.get)
                else:
                    best_algo = max(valid_values, key=valid_values.get)

            # Write row
            ws.cell(row=row_idx, column=1, value=prob_name).border = thin_border
            for col_idx, algo in enumerate(algorithm_order, 2):
                val = values[algo]
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = f"{val:.4e}" if not np.isnan(val) else 'N/A'
                cell.alignment = center_align
                cell.border = thin_border
                if algo == best_algo:
                    cell.font = best_font

            row_idx += 1

        # Runtime row
        ws.cell(row=row_idx, column=1, value='Runtime (s)').border = thin_border
        best_runtime_algo = min(runtime, key=runtime.get)
        for col_idx, algo in enumerate(algorithm_order, 2):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = f"{runtime[algo]:.2f}"
            cell.alignment = center_align
            cell.border = thin_border
            if algo == best_runtime_algo:
                cell.font = best_font

        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        for col_idx in range(2, len(algorithm_order) + 2):
            ws.column_dimensions[chr(64 + col_idx)].width = 15

        # Save file
        output_file = save_dir / 'test_results.xlsx'
        wb.save(output_file)
        print(f"Excel table saved to: {output_file}")

        return output_file


# =============================================================================
# Main Test Data Analyzer Class
# =============================================================================

class TestDataAnalyzer:
    """
    Main class for analyzing single-run test data.

    This class handles pickle files stored directly in the data folder,
    providing a lightweight analysis pipeline without statistical analysis.

    Attributes:
        data_path: Path
            Path to the data directory containing pickle files.
        settings: Optional[Dict[str, Any]]
            Problem settings including reference definitions.
        algorithm_order: Optional[List[str]]
            Custom ordering of algorithms for display.
        plot_config: PlotConfig
            Configuration for plot generation.
    """

    def __init__(
            self,
            data_path: Union[str, Path] = './Data',
            settings: Optional[Dict[str, Any]] = None,
            algorithm_order: Optional[List[str]] = None,
            save_path: Union[str, Path] = './Results',
            figure_format: str = 'pdf',
            log_scale: bool = False,
            show_pf: bool = True,
            show_nd: bool = True,
            best_so_far: bool = True,
            clear_results: bool = True,
            file_suffix: str = '.pkl'
    ):
        """
        Initialize TestDataAnalyzer.

        Parameters:
            data_path: Union[str, Path]
                Path to data directory containing pickle files.
                Default: './Data'

            settings: Optional[Dict[str, Any]]
                Problem settings dictionary for multi-objective metrics.
                Default: None (single-objective mode)

            algorithm_order: Optional[List[str]]
                Custom ordering of algorithms for display.
                Default: None (alphabetical order)

            save_path: Union[str, Path]
                Directory path to save all output files.
                Default: './Results'

            figure_format: str
                Output figure format: 'pdf', 'png', 'svg', etc.
                Default: 'pdf'

            log_scale: bool
                Whether to use logarithmic scale for convergence plot y-axis.
                Default: False

            show_pf: bool
                Whether to show true Pareto front in ND solution plots.
                Default: True

            show_nd: bool
                Whether to filter and show only non-dominated solutions.
                Default: True

            best_so_far: bool
                Whether to use best-so-far metric values.
                Default: True

            clear_results: bool
                Whether to clear existing results folder before analysis.
                Default: True

            file_suffix: str
                Suffix pattern for pickle files.
                Default: '.pkl'
        """
        self.data_path = Path(data_path)
        self.settings = settings
        self.algorithm_order = algorithm_order
        self.best_so_far = best_so_far
        self.clear_results = clear_results
        self.file_suffix = file_suffix

        self.plot_config = PlotConfig(
            figure_format=figure_format,
            log_scale=log_scale,
            show_pf=show_pf,
            show_nd=show_nd,
            save_path=Path(save_path)
        )

        self.table_save_path = Path(save_path)

        # Internal state
        self._scan_result: Optional[TestScanResult] = None
        self._metric_results: Optional[TestMetricResults] = None

    def scan_data(self) -> TestScanResult:
        """
        Scan the data directory to detect pickle files.

        Returns:
            TestScanResult
                Dataclass containing:
                    - algorithms: List[str] - Algorithm names extracted from filenames
                    - problems: List[str] - Problem names
                    - data_path: Path - Path to scanned directory
                    - file_mapping: Dict[str, Path] - Algorithm to file path mapping

        Raises:
            FileNotFoundError: If data_path does not exist.
            ValueError: If no pickle files found.
        """
        if not self.data_path.exists():
            raise FileNotFoundError(f"Data path does not exist: {self.data_path}")

        # Find all pickle files with specified suffix
        file_mapping = {}
        for pkl_file in self.data_path.glob(f'*{self.file_suffix}'):
            # Extract algorithm name from filename (remove suffix)
            algo_name = pkl_file.stem
            file_mapping[algo_name] = pkl_file

        if not file_mapping:
            raise ValueError(f"No pickle files found in {self.data_path} with suffix '{self.file_suffix}'")

        algorithms = sorted(file_mapping.keys())

        # Try to extract problem names from first file
        first_file = file_mapping[algorithms[0]]
        data = DataUtils.load_pickle(first_file)
        n_tasks = len(data['all_objs'])

        # Generate problem names
        if self.settings and 'problems' in self.settings:
            problems = self.settings['problems'][:n_tasks]
        else:
            problems = [f'Task{i + 1}' for i in range(n_tasks)]

        print(f"Found {len(algorithms)} algorithms: {algorithms}")
        print(f"Found {n_tasks} tasks/problems: {problems}")

        self._scan_result = TestScanResult(
            algorithms=algorithms,
            problems=problems,
            data_path=self.data_path,
            file_mapping=file_mapping
        )

        return self._scan_result

    def calculate_metrics(self) -> TestMetricResults:
        """
        Calculate metric values for all algorithms.

        Returns:
            TestMetricResults
                Dataclass containing all computed metrics.

        Raises:
            RuntimeError: If scan_data() has not been called.
        """
        if self._scan_result is None:
            self.scan_data()

        scan = self._scan_result
        algo_order = self.algorithm_order if self.algorithm_order else scan.algorithms
        metric_name = self.settings.get('metric') if self.settings else None

        # Initialize storage dictionaries
        all_values = {}
        all_best_values = {}
        all_objective_values = {}
        all_runtime = {}
        all_max_nfes = {}

        for algo in algo_order:
            pkl_path = scan.file_mapping[algo]
            data = DataUtils.load_pickle(pkl_path)

            metric_values, metric_values_best_so_far = self._calculate_single_algorithm_metrics(data)

            selected = metric_values_best_so_far if self.best_so_far else metric_values
            all_values[algo] = selected

            # Extract final values
            last_vals = [
                np.asarray(task_arr).ravel()[-1] if len(task_arr) > 0 else np.nan
                for task_arr in selected
            ]
            all_best_values[algo] = last_vals

            # Extract final objective values
            last_objs = [data['all_objs'][t][-1] for t in range(len(data['all_objs']))]
            all_objective_values[algo] = last_objs

            all_runtime[algo] = data['runtime']
            all_max_nfes[algo] = data['max_nfes']

            print(f"  {algo}: {len(data['all_objs'])} tasks, runtime={data['runtime']:.2f}s")

        self._metric_results = TestMetricResults(
            metric_values=all_values,
            best_values=all_best_values,
            objective_values=all_objective_values,
            runtime=all_runtime,
            max_nfes=all_max_nfes,
            metric_name=metric_name,
            problems=scan.problems
        )

        return self._metric_results

    def _calculate_single_algorithm_metrics(
            self,
            data: Dict[str, Any]
    ) -> Tuple[List[np.ndarray], List[np.ndarray]]:
        """
        Calculate metric values for a single algorithm's data.

        Parameters:
            data: Dict[str, Any]
                Loaded pickle data.

        Returns:
            Tuple[List[np.ndarray], List[np.ndarray]]
                Tuple of (metric_values, metric_values_best_so_far).
        """
        all_objs = data['all_objs']
        all_cons = data.get('all_cons', None)
        n_tasks = len(all_objs)

        metric_values = []
        metric_values_best_so_far = []

        for t in range(n_tasks):
            n_gens = len(all_objs[t])
            task_values = np.zeros(n_gens)
            task_best_so_far = np.zeros(n_gens)

            best_so_far = None

            for gen in range(n_gens):
                objs_gen = all_objs[t][gen]
                cons_gen = all_cons[t][gen] if all_cons is not None else None
                M = objs_gen.shape[1]

                if M == 1:
                    # Single-objective
                    metric_value = np.min(objs_gen[:, 0])
                    sign = -1
                else:
                    # Multi-objective
                    if self.settings is None:
                        # Default to simple dominated hypervolume approximation
                        metric_value = np.min(np.sum(objs_gen, axis=1))
                        sign = -1
                    else:
                        metric_name = self.settings.get('metric', 'IGD')
                        prob = self._scan_result.problems[t] if t < len(self._scan_result.problems) else f'P{t + 1}'
                        reference = DataUtils.load_reference(self.settings, prob, t, M)

                        if metric_name == 'IGD':
                            metric_instance = IGD()
                            metric_value = metric_instance.calculate(objs_gen, reference)
                            sign = metric_instance.sign
                        elif metric_name == 'HV':
                            metric_instance = HV()
                            # If reference is 1D or single row, treat as ref point; otherwise as PF
                            if reference.ndim == 1 or reference.shape[0] == 1:
                                ref_point = reference.flatten()
                                metric_value = metric_instance.calculate(objs_gen, reference=ref_point)
                            else:
                                metric_value = metric_instance.calculate(objs_gen, pf=reference)
                            sign = metric_instance.sign
                        elif metric_name == 'IGDp':
                            metric_instance = IGDp()
                            metric_value = metric_instance.calculate(objs_gen, reference)
                            sign = metric_instance.sign
                        elif metric_name == 'GD':
                            metric_instance = GD()
                            metric_value = metric_instance.calculate(objs_gen, reference)
                            sign = metric_instance.sign
                        elif metric_name == 'DeltaP':
                            metric_instance = DeltaP()
                            metric_value = metric_instance.calculate(objs_gen, reference)
                            sign = metric_instance.sign
                        elif metric_name == 'Spacing':
                            metric_instance = Spacing()
                            metric_value = metric_instance.calculate(objs_gen)
                            sign = metric_instance.sign
                        elif metric_name == 'Spread':
                            metric_instance = Spread()
                            metric_value = metric_instance.calculate(objs_gen, reference)
                            sign = metric_instance.sign
                        elif metric_name == 'FR':
                            if cons_gen is None:
                                raise ValueError('FR metric requires constraint data, but all_cons is not available')
                            metric_instance = FR()
                            metric_value = metric_instance.calculate(cons_gen)
                            sign = metric_instance.sign
                        elif metric_name == 'CV':
                            if cons_gen is None:
                                raise ValueError('CV metric requires constraint data, but all_cons is not available')
                            metric_instance = CV()
                            metric_value = metric_instance.calculate(cons_gen)
                            sign = metric_instance.sign
                        else:
                            raise ValueError(f'Unsupported metric: {metric_name}')

                task_values[gen] = metric_value

                if best_so_far is None:
                    best_so_far = metric_value
                else:
                    if sign == -1:
                        best_so_far = min(best_so_far, metric_value)
                    else:
                        best_so_far = max(best_so_far, metric_value)

                task_best_so_far[gen] = best_so_far

            metric_values.append(task_values)
            metric_values_best_so_far.append(task_best_so_far)

        return metric_values, metric_values_best_so_far

    def generate_convergence_plots(self) -> None:
        """Generate and save convergence curve plots."""
        if self._metric_results is None:
            self.calculate_metrics()

        algo_order = self.algorithm_order if self.algorithm_order else self._scan_result.algorithms

        plot_gen = TestPlotGenerator(self.plot_config)
        plot_gen.plot_convergence_curves(
            self._metric_results.metric_values,
            self._metric_results.max_nfes,
            algo_order,
            self._metric_results.problems,
            self._metric_results.metric_name
        )

    def generate_runtime_plots(self) -> None:
        """Generate and save runtime comparison bar plots."""
        if self._metric_results is None:
            self.calculate_metrics()

        algo_order = self.algorithm_order if self.algorithm_order else self._scan_result.algorithms

        plot_gen = TestPlotGenerator(self.plot_config)
        plot_gen.plot_runtime(self._metric_results.runtime, algo_order)

    def generate_nd_solution_plots(self) -> None:
        """Generate and save non-dominated solution plots."""
        if self._metric_results is None:
            self.calculate_metrics()

        algo_order = self.algorithm_order if self.algorithm_order else self._scan_result.algorithms

        plot_gen = TestPlotGenerator(self.plot_config)
        plot_gen.plot_nd_solutions(
            self._metric_results.objective_values,
            algo_order,
            self._metric_results.problems,
            self.settings
        )

    def generate_excel_tables(self) -> Path:
        """
        Generate Excel comparison tables.

        Returns:
            Path: Path to generated Excel file.
        """
        if self._metric_results is None:
            self.calculate_metrics()

        algo_order = self.algorithm_order if self.algorithm_order else self._scan_result.algorithms

        table_gen = TestTableGenerator(self.table_save_path)

        # Generate results table
        output_path = table_gen.generate_excel_table(
            self._metric_results.best_values,
            self._metric_results.runtime,
            algo_order,
            self._metric_results.problems,
            self._metric_results.metric_name
        )

        return output_path

    def run(self) -> TestMetricResults:
        """
        Execute the complete test analysis pipeline.

        Returns:
            TestMetricResults
                Complete metric results from the analysis.
        """
        print("=" * 60)
        print("Starting Test Data Analysis Pipeline")
        print("=" * 60)

        # Step 0: Clear results folder if requested
        if self.clear_results:
            results_path = self.plot_config.save_path
            if results_path.exists():
                print(f'\nClearing existing results folder: {results_path}')
                # Retry mechanism for Windows file locking issues
                max_retries = 3
                for attempt in range(max_retries):
                    try:
                        shutil.rmtree(results_path)
                        break
                    except PermissionError as e:
                        if attempt < max_retries - 1:
                            import time
                            print(f'  Warning: Folder in use, retrying in 1s... ({attempt + 1}/{max_retries})')
                            time.sleep(1)
                        else:
                            print(f'  Error: Cannot delete {results_path}. Please close any programs using this folder.')
                            raise e
            results_path.mkdir(parents=True, exist_ok=True)

        # Step 1: Scan data
        print('\n[1/6] Scanning test data directory...')
        self.scan_data()

        # Step 2: Calculate metrics
        print('\n[2/6] Calculating metric values...')
        self.calculate_metrics()

        # Step 3: Generate Excel tables
        print('\n[3/6] Generating Excel tables...')
        self.generate_excel_tables()

        # Step 4: Plot convergence curves
        print('\n[4/6] Plotting convergence curves...')
        self.generate_convergence_plots()

        # Step 5: Plot runtime
        print('\n[5/6] Plotting runtime comparison...')
        self.generate_runtime_plots()

        # Step 6: Plot ND solutions (if multi-objective)
        algo_order = self.algorithm_order if self.algorithm_order else self._scan_result.algorithms
        first_algo = algo_order[0]
        first_objs = self._metric_results.objective_values[first_algo][0]
        if first_objs.shape[1] > 1:
            print('\n[6/6] Plotting non-dominated solutions...')
            self.generate_nd_solution_plots()
        else:
            print('\n[6/6] Skipping ND plots (single-objective)')

        print("\n" + "=" * 60)
        print("Test Data Analysis Completed")
        print("=" * 60)

        return self._metric_results


# =============================================================================
# Module Entry Point
# =============================================================================

if __name__ == '__main__':
    """
    Usage Examples for TestDataAnalyzer
    ====================================

    Example 1: Quick Start
    ----------------------
    Analyze pickle files in ./Data folder:

        from ddmtolab.Methods.test_data_analysis import TestDataAnalyzer

        analyzer = TestDataAnalyzer(data_path='./Data')
        results = analyzer.run()


    Example 2: With Custom Settings (Multi-Objective)
    -------------------------------------------------
    Multi-objective analysis with IGD metric:

        SETTINGS = {
            'metric': 'IGD',
            'ref_path': './MOReference',
            'problems': ['ZDT1', 'ZDT2'],
            'ZDT1': {'T1': 'ZDT1_ref.npy'},
            'ZDT2': {'T1': 'ZDT2_ref.npy'}
        }

        analyzer = TestDataAnalyzer(
            data_path='./Data',
            settings=SETTINGS,
            save_path='./Results',
            figure_format='png'
        )
        results = analyzer.run()


    Example 3: Custom Algorithm Order
    ---------------------------------
        analyzer = TestDataAnalyzer(
            data_path='./Data',
            algorithm_order=['DE', 'GA', 'PSO'],
            figure_format='pdf'
        )
        results = analyzer.run()


    Example 4: Step-by-Step Analysis
    --------------------------------
        analyzer = TestDataAnalyzer(data_path='./Data')

        # Scan files
        scan_result = analyzer.scan_data()
        print(f"Found: {scan_result.algorithms}")

        # Calculate metrics
        results = analyzer.calculate_metrics()

        # Generate specific outputs
        analyzer.generate_latex_tables()
        analyzer.generate_convergence_plots()


    Example 5: Plot Customization
    -----------------------------
    Control figure format, log scale, and Pareto front display:

        analyzer = TestDataAnalyzer(
            data_path='./Data',
            figure_format='pdf',    # Output format: 'pdf', 'png', 'svg'
            log_scale=True,         # Use log scale for y-axis
            show_pf=True,           # Show true Pareto front
            show_nd=True            # Filter non-dominated solutions
        )


    Expected File Structure
    -----------------------
        ./Data/
        ├── GA.pkl
        ├── DE.pkl
        ├── PSO.pkl
        └── ...

    Each .pkl file should contain:
        - 'all_objs': List[List[np.ndarray]] - Objectives per task per generation
        - 'runtime': float - Total runtime in seconds
        - 'max_nfes': List[int] - Max function evaluations per task
    """

    # Demo run
    print("TestDataAnalyzer - Demo")
    print("=" * 50)

    analyzer = TestDataAnalyzer(
        data_path='./Data',
        save_path='./Results',
        figure_format='pdf',
        clear_results=True
    )

    # Run complete analysis
    try:
        results = analyzer.run()
    except (FileNotFoundError, ValueError) as e:
        print(f"Demo skipped: {e}")
        print("Create pickle files in ./Data/ to run analysis.")
